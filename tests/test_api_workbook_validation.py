from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from scripts.build_api_config_from_analysis import build_api_config
from scripts.export_api_spec import export_api_workbook
from scripts.validate_api_workbook import validate_api_workbook


FIXTURE_ROOT = Path(__file__).resolve().parent / "fixtures" / "phase1"
PACKAGE_ROOT = Path(__file__).resolve().parents[1]


def test_validate_api_workbook_reports_required_sheets(tmp_path: Path) -> None:
    analysis_payload = json.loads((FIXTURE_ROOT / "analysis_show.json").read_text(encoding="utf-8"))
    project_config = json.loads((PACKAGE_ROOT / "configs" / "project_config.json").read_text(encoding="utf-8"))
    api_config = build_api_config(analysis_payload, project_config)
    api_config_path = tmp_path / "api_config.json"
    api_config_path.write_text(json.dumps(api_config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    workbook_path = tmp_path / "api_spec.xlsx"
    export_api_workbook(str(api_config_path), output_path=str(workbook_path))

    report = validate_api_workbook(
        workbook_path,
        api_config=api_config,
        analysis_payload=analysis_payload,
    )

    assert report["status"] in {"pass", "review"}
    assert "表紙" in report["sheet_names"]
    assert "処理詳細" in report["sheet_names"]


def test_gaibudatatorikomi_show_workbook_regression_guards(tmp_path: Path) -> None:
    project_config = json.loads((PACKAGE_ROOT / "configs" / "project_config.json").read_text(encoding="utf-8"))
    analysis_payload = {
        "feature": "/api/gaiBuDataTorikomi/show",
        "request_params": [{"name": "baseDate"}, {"name": "functionId"}, {"name": "page"}],
        "response_params": [{"name": "kinoPermissionMap"}, {"name": "pageSize"}, {"name": "condition"}],
    }
    api_config = build_api_config(
        analysis_payload,
        project_config,
        overrides={
            "cover": {
                "create_date": "2026-04-08",
                "update_date": "2026-04-08",
            }
        },
    )
    api_config_path = tmp_path / "gaibudatatorikomi_show_api_config.json"
    api_config_path.write_text(json.dumps(api_config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    workbook_path = tmp_path / "gaibudatatorikomi_show.xlsx"

    export_api_workbook(str(api_config_path), output_path=str(workbook_path))

    wb = load_workbook(workbook_path)
    seq_ws = wb["APIシーケンス&DFD"]
    assert seq_ws["H13"].value == "サブ機能(subfunctions)"
    assert seq_ws["H16"].value == "権限グループ(role_groups)"
    assert seq_ws["H20"].value == "給与抽出(kyuyo_select)"

    detail_ws = wb["処理詳細"]
    assert detail_ws["C9"].alignment.shrink_to_fit is None
    assert detail_ws["C13"].alignment.shrink_to_fit is None
    assert detail_ws["C13"].value == "共通コンテキストを補完する"
    detail_texts = {
        cell.value
        for row in detail_ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value
    }
    assert "体系項目一覧を取得する。" in detail_texts
    assert "taikeiItemList を取得する。" not in detail_texts

    response_ws = wb["レスポンスAPIパラメーター"]
    assert response_ws.row_dimensions[33].height is None
    assert response_ws.row_dimensions[77].height is None
    for cell_ref in ("K77", "L77", "K78", "L78"):
        assert response_ws[cell_ref].border.left.style == "thin"
        assert response_ws[cell_ref].border.right.style == "thin"
    wb.close()
