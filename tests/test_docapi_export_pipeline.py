from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
FIXTURE_ROOT = Path(__file__).resolve().parent / "fixtures" / "phase1"
BACK_ROOT = FIXTURE_ROOT / "back" / "src" / "main" / "java"
CONTROLLER_DIR = BACK_ROOT / "jp" / "co" / "fminc" / "socia" / "aplAprList"


def _cli_command() -> list[str]:
    return [sys.executable, "-m", "scripts.docapi_cli"]


def _run_docapi(*args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [*_cli_command(), *args],
        cwd=PACKAGE_ROOT,
        capture_output=True,
        text=True,
        check=False,
    )


def test_docapi_generate_exports_workbook_and_manifest_artifacts(tmp_path: Path) -> None:
    result = _run_docapi(
        "generate",
        "--path",
        str(CONTROLLER_DIR),
        "--pick",
        "1",
        "--yes",
        "--non-interactive",
        "--output-dir",
        str(tmp_path),
    )

    assert result.returncode == 0, result.stderr
    run_dirs = sorted(path for path in tmp_path.iterdir() if path.is_dir())
    assert len(run_dirs) == 1
    run_dir = run_dirs[0]

    workbook_path = run_dir / "api_spec.xlsx"
    export_path = run_dir / "export.json"
    manifest_path = run_dir / "manifest.json"

    assert workbook_path.exists()
    assert export_path.exists()
    assert manifest_path.exists()

    workbook = load_workbook(workbook_path)
    assert "表紙" in workbook.sheetnames
    assert "処理概要" in workbook.sheetnames

    export_payload = json.loads(export_path.read_text(encoding="utf-8"))
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    assert manifest["artifacts"]["api_spec"] == "api_spec.xlsx"
    assert manifest["artifacts"]["export"] == "export.json"
    assert {"output_path", "template_path", "project_config_path"} <= set(export_payload)
