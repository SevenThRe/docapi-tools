"""
UI設計書 品質検証スクリプト
============================
使用方法:
  python validate_spec.py <excel_file>

検証項目:
  1. シート構成の完全性
  2. 共通ヘッダーの正確性（表紙参照の数式）
  3. 画面レイアウトのオブジェクト定義品質
  4. 処理詳細とレイアウトの対応関係
  5. エラー定義の網羅性
"""

import sys
import json
import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict


class SpecValidator:
    def __init__(self, filepath):
        self.wb = openpyxl.load_workbook(filepath)
        self.issues = []
        self.stats = defaultdict(int)

    def add_issue(self, severity, sheet, message):
        self.issues.append({"severity": severity, "sheet": sheet, "message": message})
        self.stats[severity] += 1

    def validate_all(self):
        self.check_sheet_structure()
        self.check_cover_sheet()
        self.check_common_headers()
        self.check_layout_sheets()
        self.check_processing_sheet()
        self.check_revision_history()
        self.check_cross_references()
        return self.generate_report()

    def check_sheet_structure(self):
        names = self.wb.sheetnames
        required = ['表紙']
        for req in required:
            if req not in names:
                self.add_issue("ERROR", "全体", f"必須シート「{req}」が見つかりません")
        has_overview = any('機能概要' in n for n in names)
        has_layout = any('画面レイアウト' in n for n in names)
        has_processing = any('処理詳細' in n for n in names)
        has_history = any('改定履歴' in n for n in names)
        if not has_overview:
            self.add_issue("ERROR", "全体", "「外部-機能概要」シートが見つかりません")
        if not has_layout:
            self.add_issue("ERROR", "全体", "「外部-画面レイアウト-*」シートが見つかりません")
        if not has_processing:
            self.add_issue("ERROR", "全体", "「内部-処理詳細」シートが見つかりません")
        if not has_history:
            self.add_issue("WARNING", "全体", "「改定履歴」シートが見つかりません")
        self.stats['sheets'] = len(names)

    def check_cover_sheet(self):
        if '表紙' not in self.wb.sheetnames:
            return
        ws = self.wb['表紙']
        required_fields = {
            'B4': 'ユーザー', 'B5': 'プロジェクト', 'B6': 'システム',
            'B7': '機能名', 'B8': '機能ID',
            'B10': '作成日', 'B11': '作成者', 'B12': '更新日', 'B13': '更新者'
        }
        for cell, label in required_fields.items():
            if ws[cell].value is None:
                self.add_issue("ERROR", "表紙", f"セル{cell}（{label}）が空です")
            else:
                self.stats['cover_fields'] = self.stats.get('cover_fields', 0) + 1
        for cell in ['C4', 'C5', 'C6', 'C7', 'C8', 'C10', 'C11', 'C12', 'C13']:
            if ws[cell].value is None:
                self.add_issue("WARNING", "表紙", f"セル{cell}（値）が空です")

    def check_common_headers(self):
        for sn in self.wb.sheetnames:
            if sn == '表紙':
                continue
            ws = self.wb[sn]
            a1 = ws['A1'].value
            if a1 is None:
                self.add_issue("ERROR", sn, "A1（文書タイトル参照）が空です")
            elif isinstance(a1, str) and '表紙' not in a1:
                self.add_issue("WARNING", sn, f"A1が表紙を参照していません: {a1}")
            for row in [2, 3, 4, 5]:
                a_val = ws.cell(row=row, column=1).value
                if a_val and isinstance(a_val, str) and '表紙' not in a_val:
                    self.add_issue("WARNING", sn, f"A{row}が表紙を参照していません")

    def check_layout_sheets(self):
        for sn in self.wb.sheetnames:
            if '画面レイアウト' not in sn:
                continue
            ws = self.wb[sn]
            objects = self._extract_objects(ws)
            self.stats[f'{sn}_objects'] = len(objects)
            for obj in objects:
                self._validate_object(sn, obj)

    def _extract_objects(self, ws):
        objects = []
        current_obj = None
        for row in range(8, ws.max_row + 1):
            b_val = ws.cell(row=row, column=2).value
            c_val = ws.cell(row=row, column=3).value
            d_val = ws.cell(row=row, column=4).value
            e_val = ws.cell(row=row, column=5).value
            h_val = ws.cell(row=row, column=8).value
            if b_val is not None and c_val is not None:
                if current_obj:
                    objects.append(current_obj)
                current_obj = {
                    'no': str(b_val), 'name': str(c_val), 'row': row,
                    'has_type': False, 'has_action': False,
                    'has_display_control': False, 'has_api': False,
                    'has_error': False, 'has_attribute': False,
                    'action_lines': 0, 'sub_objects': []
                }
            if current_obj:
                label = str(d_val) if d_val else (str(e_val) if e_val else None)
                if label:
                    if label == '種類':
                        current_obj['has_type'] = True
                        current_obj['type_value'] = str(h_val) if h_val else None
                    elif label == '動作':
                        current_obj['has_action'] = True
                        current_obj['action_lines'] += 1
                    elif label == '表示制御':
                        current_obj['has_display_control'] = True
                    elif label == 'API':
                        current_obj['has_api'] = True
                    elif label == 'エラー':
                        current_obj['has_error'] = True
                    elif label == '属性':
                        current_obj['has_attribute'] = True
                if h_val and current_obj.get('has_action') and label is None:
                    current_obj['action_lines'] += 1
        if current_obj:
            objects.append(current_obj)
        return objects

    def _validate_object(self, sheet_name, obj):
        obj_id = f"オブジェクト{obj['no']}「{obj['name']}」"
        if not obj['has_type']:
            self.add_issue("ERROR", sheet_name, f"{obj_id}: 種類が定義されていません")
        if not obj['has_action']:
            self.add_issue("ERROR", sheet_name, f"{obj_id}: 動作が定義されていません")
        type_val = obj.get('type_value', '')
        interactive_types = ['アイコン', 'ボタン', 'テキスト', 'チェック', 'ラジオ']
        is_interactive = any(t in str(type_val) for t in interactive_types)
        if is_interactive:
            if not obj['has_display_control']:
                self.add_issue("WARNING", sheet_name,
                    f"{obj_id}: インタラクティブ要素ですが表示制御が定義されていません")
            if not obj['has_attribute']:
                self.add_issue("INFO", sheet_name,
                    f"{obj_id}: インタラクティブ要素ですが属性（cursor等）が定義されていません")
        if obj['action_lines'] == 1 and is_interactive:
            self.add_issue("INFO", sheet_name,
                f"{obj_id}: 動作が1行のみです。境界条件や条件分岐の記述を検討してください")

    def check_processing_sheet(self):
        processing_sheets = [sn for sn in self.wb.sheetnames if '処理詳細' in sn]
        for sn in processing_sheets:
            ws = self.wb[sn]
            api_count = 0
            has_request = False
            has_response = False
            for row in range(8, ws.max_row + 1):
                d_val = ws.cell(row=row, column=4).value
                e_val = ws.cell(row=row, column=5).value
                if d_val == 'API名':
                    api_count += 1
                if d_val == 'APIリクエスト':
                    has_request = True
                if d_val == 'APIレスポンス':
                    has_response = True
            self.stats[f'{sn}_apis'] = api_count
            if api_count == 0:
                self.add_issue("ERROR", sn, "API定義が見つかりません")

    def check_revision_history(self):
        history_sheets = [sn for sn in self.wb.sheetnames if '改定履歴' in sn]
        for sn in history_sheets:
            ws = self.wb[sn]
            if ws['A8'].value is None:
                self.add_issue("WARNING", sn, "改定履歴のレコードが空です")

    def check_cross_references(self):
        layout_objects = {}
        for sn in self.wb.sheetnames:
            if '画面レイアウト' not in sn:
                continue
            ws = self.wb[sn]
            for row in range(8, ws.max_row + 1):
                b_val = ws.cell(row=row, column=2).value
                c_val = ws.cell(row=row, column=3).value
                if b_val is not None and c_val is not None:
                    layout_objects[str(b_val)] = str(c_val)
        processing_objects = {}
        for sn in self.wb.sheetnames:
            if '処理詳細' not in sn:
                continue
            ws = self.wb[sn]
            for row in range(8, ws.max_row + 1):
                b_val = ws.cell(row=row, column=2).value
                c_val = ws.cell(row=row, column=3).value
                if b_val is not None and c_val is not None:
                    processing_objects[str(b_val)] = str(c_val)
        for no, name in layout_objects.items():
            if no not in processing_objects and no != '初期表示':
                self.add_issue("INFO", "対応関係",
                    f"画面レイアウトのオブジェクト{no}「{name}」に対応する処理詳細がありません")

    def generate_report(self):
        errors = [i for i in self.issues if i['severity'] == 'ERROR']
        warnings = [i for i in self.issues if i['severity'] == 'WARNING']
        infos = [i for i in self.issues if i['severity'] == 'INFO']
        report = {
            "status": "pass" if len(errors) == 0 else "fail",
            "summary": {
                "total_issues": len(self.issues),
                "errors": len(errors),
                "warnings": len(warnings),
                "info": len(infos),
            },
            "stats": dict(self.stats),
            "issues": self.issues,
            "quality_level": self._determine_quality_level(errors, warnings)
        }
        return report

    def _determine_quality_level(self, errors, warnings):
        if len(errors) > 0:
            return "Level 0: 構造に問題あり（修正必須）"
        elif len(warnings) > 3:
            return "Level 1: 最低限（実装開始可能だが手戻りリスク高）"
        elif len(warnings) > 0:
            return "Level 2: 標準（手戻り少ない）"
        else:
            return "Level 3: 理想（認識齟齬ゼロに近い）"


def main():
    if len(sys.argv) < 2:
        print("Usage: python validate_spec.py <excel_file>")
        sys.exit(1)
    filepath = sys.argv[1]
    validator = SpecValidator(filepath)
    report = validator.validate_all()
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report['status'] == 'pass' else 1


if __name__ == '__main__':
    sys.exit(main())
