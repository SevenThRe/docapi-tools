#!/usr/bin/env python3
"""
Web UI設計書ジェネレーター
========================

Excelフォーマット仕様に従ったUI設計書(.xlsx)を自動生成するスクリプト。

使用方法:
    python generate_spec.py config.json
    python generate_spec.py --config config.json --output output.xlsx

設定フォーマット (JSON):
    {
        "cover": {
            "company": "株式会社エフエム",
            "project": "Socia2026",
            "system": "Sociaポータル",
            "function_name": "(共通)現職エリア",
            "function_id": "gensyokuArea",
            "author": "ISI李",
            "create_date": "2024-11-15",
            "update_date": "2026-01-19",
            "update_author": "ISI孫"
        },
        "overview": {
            "content": "この機能は..."
        },
        "screens": [
            {
                "name": "現職エリア",
                "objects": [
                    {
                        "no": 1,
                        "name": "エリアタイトル",
                        "type": "ラベル",
                        "attributes": [],
                        "actions": [],
                        "display_controls": [],
                        "error_definitions": []
                    }
                ]
            }
        ],
        "processing": {
            "apis": []
        }
    }
"""

import json
import sys
import argparse
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================================
# Constants
# ============================================================================

# Font styles
FONT_COVER_TITLE = Font(name='Kosugi', size=36)
FONT_COVER_LABEL = Font(name='Kosugi', size=14)
FONT_SHEET_TITLE = Font(name='Kosugi', size=20)
FONT_BODY = Font(name='Kosugi', size=9)
FONT_HEADER_VAL = Font(name='Kosugi', size=8)
FONT_COPYRIGHT = Font(name='Kosugi', size=8)

# Fill styles
FILL_HEADER = PatternFill(fill_type='solid', fgColor='F2F2F2')

# Border styles
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
BORDER_TB = Border(
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
BORDER_T = Border(top=Side(style='thin'))
BORDER_B = Border(bottom=Side(style='thin'))
BORDER_L = Border(left=Side(style='thin'))
BORDER_R = Border(right=Side(style='thin'))
BORDER_LR = Border(left=Side(style='thin'), right=Side(style='thin'))
BORDER_TB_L = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))
BORDER_TB_R = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))

# Alignment
ALIGN_CV = Alignment(vertical='center')
ALIGN_LV = Alignment(horizontal='left', vertical='center')
ALIGN_CC = Alignment(horizontal='center', vertical='center')
ALIGN_CCW = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_R = Alignment(horizontal='right')
ALIGN_LVW = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Column widths - EXACT from original file
# 表紙 sheet
COVER_COL_WIDTHS = {
    'A': 19.1796875,
    'B': 27.0,
    'C': 112.81640625,
    'D': 19.1796875,
    'E': 9.36328125
}

# 外部-機能概要 and 改定履歴 sheets (narrow format)
NARROW_COL_WIDTHS = {
    'A': 10.1796875,
    'B': 14.453125,
    'C': 2.81640625,
    'H': 111.6328125,
    'I': 3.1796875,
    'J': 11.81640625,
    'K': 11.81640625,
    'L': 9.1796875
}

# 外部-画面レイアウト sheets and 内部-処理詳細 (wide format)
# These sheets have SPECIFIC column widths set, others use DEFAULT (2.63)
WIDE_EXPLICIT_COL_WIDTHS = {
    'A': 10.1796875,
    'B': 14.453125,
    'C': 2.6328125,
    'AC': 2.6328125,
    'AK': 2.6328125,
    'AX': 2.6328125,
    'AY': 11.81640625,
    'AZ': 11.81640625
}

# Row heights
COVER_ROW_HEIGHTS = {
    1: 124.5, 2: 79.5, 3: 99.75, 4: 23.15, 5: 23.25, 6: 23.25, 7: 23.25,
    8: 23.25, 9: 16.3, 10: 23.15, 11: 23.15, 12: 23.15, 13: 23.15, 14: 63.75, 15: 21.0
}

HEADER_ROW_HEIGHTS = {
    1: 12.0, 2: 12.0, 3: 12.0, 4: 12.0, 5: 12.0, 6: 4.5
}


# ============================================================================
# SpecGenerator Class
# ============================================================================

class SpecGenerator:
    """UI設計書Excelジェネレーターのメインクラス"""

    def __init__(self, config_dict):
        """
        Args:
            config_dict (dict): 設定ディクショナリ
        """
        self.config = config_dict
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        self._validate_config()

    def _validate_config(self):
        """設定の必須項目をチェック"""
        required_top = ['cover', 'screens']
        for key in required_top:
            if key not in self.config:
                raise ValueError(f"設定に必須キー '{key}' がありません")

        cover = self.config['cover']
        required_cover = ['company', 'project', 'system', 'function_name', 'function_id',
                         'author', 'create_date', 'update_date', 'update_author']
        for key in required_cover:
            if key not in cover:
                raise ValueError(f"cover設定に必須キー '{key}' がありません")

    def _set_cover_column_widths(self, ws):
        """表紙シートの列幅を設定"""
        for col, width in COVER_COL_WIDTHS.items():
            ws.column_dimensions[col].width = width

    def _set_cover_row_heights(self, ws):
        """表紙シートの行高を設定"""
        for row, height in COVER_ROW_HEIGHTS.items():
            ws.row_dimensions[row].height = height

    def _set_narrow_column_widths(self, ws):
        """狭幅シート（機能概要、改定履歴）の列幅を設定"""
        for col, width in NARROW_COL_WIDTHS.items():
            ws.column_dimensions[col].width = width

    def _set_wide_column_widths(self, ws):
        """広幅シート（画面レイアウト、処理詳細）の列幅を設定

        EXACT specification from original file:
        - A: 10.1796875
        - B: 14.453125
        - C: 2.6328125
        - D-AW: DEFAULT width (approximately 2.63)
        - AX: 2.6328125
        - AY: 11.81640625
        - AZ: 11.81640625
        """
        # Set explicit widths
        for col, width in WIDE_EXPLICIT_COL_WIDTHS.items():
            ws.column_dimensions[col].width = width

    def _set_header_row_heights(self, ws):
        """共通ヘッダー部分の行高を設定"""
        for row, height in HEADER_ROW_HEIGHTS.items():
            ws.row_dimensions[row].height = height

    def _setup_narrow_header(self, ws, sheet_title, sub_label, desc_label):
        """狭幅シートの共通ヘッダーをセットアップ（機能概要、改定履歴用）"""
        # 列幅の設定
        self._set_narrow_column_widths(ws)
        self._set_header_row_heights(ws)

        # 行1: 左上タイトル＋シートタイトル＋版管理
        # A1 = =表紙!A2
        ws['A1'] = '=表紙!A2'
        ws['A1'].font = FONT_BODY
        ws['A1'].fill = FILL_HEADER
        ws['A1'].border = BORDER_THIN
        ws['A1'].alignment = ALIGN_CC
        ws.merge_cells('A1:G1')

        # H1:H5 結合 = シートタイトル
        ws['H1'] = sheet_title
        ws['H1'].font = FONT_SHEET_TITLE
        ws['H1'].alignment = ALIGN_CC
        ws.merge_cells('H1:H5')

        # I1:K1 ヘッダー (版管理)
        ws['I1'] = '版'
        ws['J1'] = '年月日'
        ws['K1'] = '作成/更新者'
        for col in ['I', 'J', 'K']:
            ws[f'{col}1'].font = FONT_HEADER_VAL
            ws[f'{col}1'].fill = FILL_HEADER
            ws[f'{col}1'].alignment = ALIGN_CC
            ws[f'{col}1'].border = BORDER_THIN

        # 行2: プロジェクト情報 + 版管理初版
        ws['A2'] = '=表紙!B4'
        ws['B2'] = '=表紙!C4'
        for col in ['A', 'B']:
            ws[f'{col}2'].font = FONT_HEADER_VAL
            ws[f'{col}2'].border = BORDER_THIN
        ws.merge_cells('B2:G2')

        ws['I2'] = '初'
        ws['J2'] = '=表紙!C10'
        ws['K2'] = '=表紙!C11'
        for col in ['I', 'J', 'K']:
            ws[f'{col}2'].font = FONT_HEADER_VAL
            ws[f'{col}2'].border = BORDER_THIN

        # 行3: プロジェクト
        ws['A3'] = '=表紙!B5'
        ws['B3'] = '=表紙!C5'
        for col in ['A', 'B']:
            ws[f'{col}3'].font = FONT_HEADER_VAL
            ws[f'{col}3'].border = BORDER_THIN
        ws.merge_cells('B3:G3')

        # 行4: システム
        ws['A4'] = '=表紙!B6'
        ws['B4'] = '=表紙!C6'
        for col in ['A', 'B']:
            ws[f'{col}4'].font = FONT_HEADER_VAL
            ws[f'{col}4'].border = BORDER_THIN
        ws.merge_cells('B4:G4')

        # 行5: 機能名 + 最終版
        ws['A5'] = '=表紙!B7'
        ws['B5'] = '=表紙!C7'
        for col in ['A', 'B']:
            ws[f'{col}5'].font = FONT_HEADER_VAL
            ws[f'{col}5'].border = BORDER_THIN
        ws.merge_cells('B5:G5')

        ws['I5'] = '最終'
        ws['J5'] = '=表紙!C12'
        ws['K5'] = '=表紙!C13'
        for col in ['I', 'J', 'K']:
            ws[f'{col}5'].font = FONT_HEADER_VAL
            ws[f'{col}5'].border = BORDER_THIN

        # 行6: 区切り線
        ws.row_dimensions[6].height = 4.5
        ws['A6'].border = BORDER_TB
        ws.merge_cells('A6:K6')

        # 行7: トラッキングバー
        ws['A7'] = 'チケットNo.'
        ws['B7'] = sub_label
        ws['C7'] = desc_label
        ws['J7'] = '改定日付'
        ws['K7'] = '改定者'
        ws.merge_cells('C7:H7')

        for col in ['A', 'B', 'C', 'J', 'K']:
            ws[f'{col}7'].font = FONT_HEADER_VAL
            ws[f'{col}7'].fill = FILL_HEADER
            ws[f'{col}7'].border = BORDER_THIN
            ws[f'{col}7'].alignment = ALIGN_CC

    def _setup_wide_header(self, ws, sheet_title, sub_label, desc_label):
        """広幅シートの共通ヘッダーをセットアップ（画面レイアウト、処理詳細用）"""
        # 列幅の設定
        self._set_wide_column_widths(ws)
        self._set_header_row_heights(ws)

        # 行1: 左上タイトル＋シートタイトル＋版管理
        ws['A1'] = '=表紙!A2'
        ws['A1'].font = FONT_BODY
        ws['A1'].fill = FILL_HEADER
        ws['A1'].border = BORDER_THIN
        ws['A1'].alignment = ALIGN_CC
        ws.merge_cells('A1:G1')

        # H1:AW5 結合 = シートタイトル
        ws['H1'] = sheet_title
        ws['H1'].font = FONT_SHEET_TITLE
        ws['H1'].alignment = ALIGN_CC
        ws.merge_cells('H1:AW5')

        # AX1:AZ1 ヘッダー (版管理)
        ws['AX1'] = '版'
        ws['AY1'] = '年月日'
        ws['AZ1'] = '作成/更新者'
        for col in ['AX', 'AY', 'AZ']:
            ws[f'{col}1'].font = FONT_HEADER_VAL
            ws[f'{col}1'].fill = FILL_HEADER
            ws[f'{col}1'].alignment = ALIGN_CC
            ws[f'{col}1'].border = BORDER_THIN

        # 行2: プロジェクト情報 + 版管理初版
        ws['A2'] = '=表紙!B4'
        ws['B2'] = '=表紙!C4'
        for col in ['A', 'B']:
            ws[f'{col}2'].font = FONT_HEADER_VAL
            ws[f'{col}2'].border = BORDER_THIN
        ws.merge_cells('B2:G2')

        ws['AX2'] = '初'
        ws['AY2'] = '=表紙!C10'
        ws['AZ2'] = '=表紙!C11'
        for col in ['AX', 'AY', 'AZ']:
            ws[f'{col}2'].font = FONT_HEADER_VAL
            ws[f'{col}2'].border = BORDER_THIN

        # 行3: プロジェクト
        ws['A3'] = '=表紙!B5'
        ws['B3'] = '=表紙!C5'
        for col in ['A', 'B']:
            ws[f'{col}3'].font = FONT_HEADER_VAL
            ws[f'{col}3'].border = BORDER_THIN
        ws.merge_cells('B3:G3')

        # 行4: システム
        ws['A4'] = '=表紙!B6'
        ws['B4'] = '=表紙!C6'
        for col in ['A', 'B']:
            ws[f'{col}4'].font = FONT_HEADER_VAL
            ws[f'{col}4'].border = BORDER_THIN
        ws.merge_cells('B4:G4')

        # 行5: 機能名 + 最終版
        ws['A5'] = '=表紙!B7'
        ws['B5'] = '=表紙!C7'
        for col in ['A', 'B']:
            ws[f'{col}5'].font = FONT_HEADER_VAL
            ws[f'{col}5'].border = BORDER_THIN
        ws.merge_cells('B5:G5')

        ws['AX5'] = '最終'
        ws['AY5'] = '=表紙!C12'
        ws['AZ5'] = '=表紙!C13'
        for col in ['AX', 'AY', 'AZ']:
            ws[f'{col}5'].font = FONT_HEADER_VAL
            ws[f'{col}5'].border = BORDER_THIN

        # 行6: 区切り線
        ws.row_dimensions[6].height = 4.5
        ws['A6'].border = BORDER_TB
        ws.merge_cells('A6:AZ6')

        # 行7: トラッキングバー
        ws['A7'] = 'チケットNo.'
        ws['B7'] = sub_label
        ws['C7'] = desc_label
        ws['AY7'] = '改定日付'
        ws['AZ7'] = '改定者'
        ws.merge_cells('C7:AX7')

        for col in ['A', 'B', 'C', 'AY', 'AZ']:
            ws[f'{col}7'].font = FONT_HEADER_VAL
            ws[f'{col}7'].fill = FILL_HEADER
            ws[f'{col}7'].border = BORDER_THIN
            ws[f'{col}7'].alignment = ALIGN_CC

    def generate_cover_sheet(self):
        """表紙シートを生成"""
        ws = self.wb.create_sheet('表紙', 0)
        self._set_cover_column_widths(ws)
        self._set_cover_row_heights(ws)

        cover = self.config['cover']

        # A2: UI設計書（36pt, center）
        ws['A2'] = 'UI設計書'
        ws['A2'].font = FONT_COVER_TITLE
        ws['A2'].alignment = ALIGN_CC

        # プロジェクト情報（B4-C8）
        ws['B4'] = 'ユーザー'
        ws['C4'] = cover['company']
        ws['B5'] = 'プロジェクト'
        ws['C5'] = cover['project']
        ws['B6'] = 'システム'
        ws['C6'] = cover['system']
        ws['B7'] = '機能名'
        ws['C7'] = cover['function_name']
        ws['B8'] = '機能ID'
        ws['C8'] = cover['function_id']

        for row in [4, 5, 6, 7, 8]:
            ws[f'B{row}'].font = FONT_COVER_LABEL
            ws[f'C{row}'].font = FONT_COVER_LABEL

        # 日付情報（B10-C13）
        # Convert date strings to datetime if needed
        create_date = cover['create_date']
        update_date = cover['update_date']
        if isinstance(create_date, str):
            try:
                create_date = datetime.strptime(create_date, '%Y-%m-%d')
            except:
                create_date = datetime.now()
        if isinstance(update_date, str):
            try:
                update_date = datetime.strptime(update_date, '%Y-%m-%d')
            except:
                update_date = datetime.now()

        ws['B10'] = '作成日'
        ws['C10'] = create_date
        ws['B11'] = '作成者'
        ws['C11'] = cover['author']
        ws['B12'] = '更新日'
        ws['C12'] = update_date
        ws['B13'] = '更新者'
        ws['C13'] = cover['update_author']

        for row in [10, 11, 12, 13]:
            ws[f'B{row}'].font = FONT_COVER_LABEL
            ws[f'C{row}'].font = FONT_COVER_LABEL

        # Copyright line (D15)
        ws['D15'] = f"©{cover['company']}"
        ws['D15'].font = FONT_COPYRIGHT
        ws['D15'].alignment = ALIGN_R

    def generate_overview_sheet(self):
        """機能概要シートを生成"""
        ws = self.wb.create_sheet('外部-機能概要')
        self._setup_narrow_header(ws, '機能概要', 'オブジェクトNo.', '説明')

        # Content placeholder (rows 8+)
        overview = self.config.get('overview', {})
        content = overview.get('content', '（機能概要を入力してください）')

        if content:
            ws['H8'] = content
            ws['H8'].font = FONT_BODY
            ws['H8'].alignment = ALIGN_LVW

    def generate_screen_layout_sheets(self):
        """画面レイアウトシートを生成（画面数分）"""
        screens = self.config.get('screens', [])

        for screen_idx, screen in enumerate(screens, 1):
            screen_name = screen['name']
            sheet_name = f'外部-画面レイアウト-{screen_name}'
            ws = self.wb.create_sheet(sheet_name)
            self._setup_wide_header(ws, f'画面レイアウト-{screen_name}', 'オブジェクトNo.', '説明')

            # オブジェクト定義（rows 8+）
            objects = screen.get('objects', [])
            current_row = 8

            for obj in objects:
                # Object no. and name in B and C columns
                ws.cell(row=current_row, column=2).value = obj.get('no')  # B列
                ws.cell(row=current_row, column=3).value = obj.get('name')  # C列

                for col in [2, 3]:
                    ws.cell(row=current_row, column=col).font = FONT_BODY
                    ws.cell(row=current_row, column=col).border = BORDER_THIN

                current_row += 1

                # Type definition (種類) - D列 with value in H列
                obj_type = obj.get('type')
                if obj_type:
                    ws.cell(row=current_row, column=4).value = '種類'  # D列
                    ws.cell(row=current_row, column=8).value = obj_type  # H列
                    for col in [4, 8]:
                        ws.cell(row=current_row, column=col).font = FONT_BODY
                        ws.cell(row=current_row, column=col).border = BORDER_THIN
                    current_row += 1

                # Attributes
                attributes = obj.get('attributes', [])
                if attributes:
                    for attr in attributes:
                        ws.cell(row=current_row, column=4).value = '属性'  # D列
                        ws.cell(row=current_row, column=8).value = attr  # H列
                        for col in [4, 8]:
                            ws.cell(row=current_row, column=col).font = FONT_BODY
                            ws.cell(row=current_row, column=col).border = BORDER_THIN
                        current_row += 1

                # Actions
                actions = obj.get('actions', [])
                if actions:
                    for action_idx, action in enumerate(actions):
                        if action_idx == 0:
                            ws.cell(row=current_row, column=4).value = '動作'  # D列 only on first action
                        ws.cell(row=current_row, column=8).value = action  # H列
                        for col in [4, 8]:
                            if ws.cell(row=current_row, column=col).value:
                                ws.cell(row=current_row, column=col).font = FONT_BODY
                                ws.cell(row=current_row, column=col).border = BORDER_THIN
                        current_row += 1

                # Display Controls
                display_controls = obj.get('display_controls', [])
                if display_controls:
                    for ctrl in display_controls:
                        ws.cell(row=current_row, column=4).value = '表示制御'  # D列
                        ws.cell(row=current_row, column=8).value = ctrl  # H列
                        for col in [4, 8]:
                            ws.cell(row=current_row, column=col).font = FONT_BODY
                            ws.cell(row=current_row, column=col).border = BORDER_THIN
                        current_row += 1

                # Error Definitions
                error_defs = obj.get('error_definitions', [])
                if error_defs:
                    for error in error_defs:
                        # Label row: D列で"エラー"
                        ws.cell(row=current_row, column=4).value = 'エラー'  # D列

                        # H:AF merged
                        cell_h = ws.cell(row=current_row, column=8)
                        cell_h.value = error.get('condition', '（エラー条件を入力してください）')
                        ws.merge_cells(f'H{current_row}:AF{current_row}')
                        ws.cell(row=current_row, column=8).fill = FILL_HEADER
                        ws.cell(row=current_row, column=8).font = FONT_BODY
                        ws.cell(row=current_row, column=8).border = BORDER_THIN

                        # AG:AQ merged
                        cell_ag = ws.cell(row=current_row, column=33)
                        cell_ag.value = error.get('message_id', 'MSG_00000')
                        ws.merge_cells(f'AG{current_row}:AQ{current_row}')
                        ws.cell(row=current_row, column=33).fill = FILL_HEADER
                        ws.cell(row=current_row, column=33).font = FONT_BODY
                        ws.cell(row=current_row, column=33).border = BORDER_THIN

                        # AR:AW merged
                        cell_ar = ws.cell(row=current_row, column=44)
                        cell_ar.value = error.get('object_no', '')
                        ws.merge_cells(f'AR{current_row}:AW{current_row}')
                        ws.cell(row=current_row, column=44).fill = FILL_HEADER
                        ws.cell(row=current_row, column=44).font = FONT_BODY
                        ws.cell(row=current_row, column=44).border = BORDER_THIN

                        current_row += 1

    def generate_processing_sheet(self):
        """処理詳細シートを生成

        EXACT hierarchy structure from original:
        Row 9:  B: Component Name (gensyokuarea)
        Row 11: C: 初期表示 (Action type)
        Row 12: D: API名
        Row 13: E: (共通)現職エリア-表示 (API method name)
        Row 14: D: APIリクエスト
        Row 15: E:Q: リクエストパラメーター [merged] | R:AT: 値 [merged]
        Row 16+: E:Q: [parameter] | R:AT: [value]
        """
        ws = self.wb.create_sheet('内部-処理詳細')
        self._setup_wide_header(ws, '処理詳細', '参照オブジェクトNo.', '処理')

        # API definitions (rows 8+)
        processing = self.config.get('processing', {})
        apis = processing.get('apis', [])
        current_row = 9

        for api in apis:
            # Row: B column = Component Name (physical name)
            ws.cell(row=current_row, column=2).value = api.get('component_name', '（コンポーネント名を入力してください）')  # B列
            ws.cell(row=current_row, column=2).font = FONT_BODY
            ws.cell(row=current_row, column=2).alignment = ALIGN_LV
            current_row += 1

            # Skip row (empty)
            current_row += 1

            # Action type row: C column = 初期表示 or other action type
            action_type = api.get('action_type', '初期表示')
            ws.cell(row=current_row, column=3).value = action_type  # C列
            ws.cell(row=current_row, column=3).font = FONT_BODY
            current_row += 1

            # API name header: D column = "API名"
            ws.cell(row=current_row, column=4).value = 'API名'  # D列
            ws.cell(row=current_row, column=4).font = FONT_BODY
            ws.cell(row=current_row, column=4).border = BORDER_THIN
            current_row += 1

            # API method name: E column
            ws.cell(row=current_row, column=5).value = api.get('name', '（API名を入力してください）')  # E列
            ws.cell(row=current_row, column=5).font = FONT_BODY
            ws.cell(row=current_row, column=5).border = BORDER_THIN
            current_row += 1

            # Request section header: D column = "APIリクエスト"
            ws.cell(row=current_row, column=4).value = 'APIリクエスト'  # D列
            ws.cell(row=current_row, column=4).font = FONT_BODY
            ws.cell(row=current_row, column=4).border = BORDER_THIN
            current_row += 1

            # Request table header
            cell_e = ws.cell(row=current_row, column=5)
            cell_e.value = 'リクエストパラメーター'
            cell_r = ws.cell(row=current_row, column=18)
            cell_r.value = '値'

            ws.merge_cells(f'E{current_row}:Q{current_row}')
            ws.cell(row=current_row, column=5).fill = FILL_HEADER
            ws.cell(row=current_row, column=5).font = FONT_HEADER_VAL
            ws.cell(row=current_row, column=5).border = BORDER_THIN
            ws.cell(row=current_row, column=5).alignment = ALIGN_CC

            ws.merge_cells(f'R{current_row}:AT{current_row}')
            ws.cell(row=current_row, column=18).fill = FILL_HEADER
            ws.cell(row=current_row, column=18).font = FONT_HEADER_VAL
            ws.cell(row=current_row, column=18).border = BORDER_THIN
            ws.cell(row=current_row, column=18).alignment = ALIGN_CC
            current_row += 1

            # Request parameters
            params = api.get('request_params', [])
            for param in params:
                cell_e = ws.cell(row=current_row, column=5)
                cell_e.value = param.get('name', '（パラメーター名を入力してください）')
                cell_r = ws.cell(row=current_row, column=18)
                cell_r.value = param.get('description', '（説明を入力してください）')

                ws.merge_cells(f'E{current_row}:Q{current_row}')
                ws.cell(row=current_row, column=5).font = FONT_BODY
                ws.cell(row=current_row, column=5).border = BORDER_THIN

                ws.merge_cells(f'R{current_row}:AT{current_row}')
                ws.cell(row=current_row, column=18).font = FONT_BODY
                ws.cell(row=current_row, column=18).border = BORDER_THIN
                current_row += 1

            # Blank row
            current_row += 1

            # Response section header: D column = "APIレスポンス"
            ws.cell(row=current_row, column=4).value = 'APIレスポンス'  # D列
            ws.cell(row=current_row, column=4).font = FONT_BODY
            ws.cell(row=current_row, column=4).border = BORDER_THIN
            current_row += 1

            # Response description
            ws.cell(row=current_row, column=5).value = api.get('response_description', '（レスポンスの処理説明を入力してください）')
            ws.cell(row=current_row, column=5).font = FONT_BODY
            ws.cell(row=current_row, column=5).border = BORDER_THIN
            current_row += 1

            # Response display mode
            ws.cell(row=current_row, column=5).value = api.get('display_mode', '■ブロック表示')
            ws.cell(row=current_row, column=5).font = FONT_BODY
            ws.cell(row=current_row, column=5).border = BORDER_THIN
            current_row += 1

            # Response table header
            cell_e = ws.cell(row=current_row, column=5)
            cell_e.value = '画面項目'
            cell_r = ws.cell(row=current_row, column=18)
            cell_r.value = 'レスポンスパラメーター'

            ws.merge_cells(f'E{current_row}:Q{current_row}')
            ws.cell(row=current_row, column=5).fill = FILL_HEADER
            ws.cell(row=current_row, column=5).font = FONT_HEADER_VAL
            ws.cell(row=current_row, column=5).border = BORDER_THIN
            ws.cell(row=current_row, column=5).alignment = ALIGN_CC

            ws.merge_cells(f'R{current_row}:AT{current_row}')
            ws.cell(row=current_row, column=18).fill = FILL_HEADER
            ws.cell(row=current_row, column=18).font = FONT_HEADER_VAL
            ws.cell(row=current_row, column=18).border = BORDER_THIN
            ws.cell(row=current_row, column=18).alignment = ALIGN_CC
            current_row += 1

            # Response parameters
            response_params = api.get('response_params', [])
            for param in response_params:
                cell_e = ws.cell(row=current_row, column=5)
                cell_e.value = param.get('screen_item', '（画面項目を入力してください）')
                cell_r = ws.cell(row=current_row, column=18)
                cell_r.value = param.get('response_param', '（レスポンスパラメーターを入力してください）')

                # Note in AA column if present (BEFORE merging)
                if 'note' in param and param['note']:
                    cell_aa = ws.cell(row=current_row, column=27)
                    cell_aa.value = f"※{param['note']}"  # AA column
                    cell_aa.font = FONT_BODY

                ws.merge_cells(f'E{current_row}:Q{current_row}')
                ws.cell(row=current_row, column=5).font = FONT_BODY
                ws.cell(row=current_row, column=5).border = BORDER_THIN

                ws.merge_cells(f'R{current_row}:AT{current_row}')
                ws.cell(row=current_row, column=18).font = FONT_BODY
                ws.cell(row=current_row, column=18).border = BORDER_THIN

                current_row += 1

            # Blank line before next API
            current_row += 1

    def generate_revision_history_sheet(self):
        """改定履歴シートを生成"""
        ws = self.wb.create_sheet('改定履歴')
        self._setup_narrow_header(ws, '改定履歴', 'シート名', '改定内容')

        # Initial revision entry (rows 8+)
        ws['A8'] = ''  # Ticket No placeholder
        ws['B8'] = '表紙'
        ws['C8'] = '初版作成'
        ws['J8'] = '=表紙!C10'
        ws['K8'] = '=表紙!C11'

        for col in ['A', 'B', 'C', 'J', 'K']:
            ws[f'{col}8'].font = FONT_BODY
            ws[f'{col}8'].border = BORDER_THIN

        # Add special formulas in row 5 for latest update
        ws['J5'] = '=MAX($J$8:$J$876)'
        ws['K5'] = '=VLOOKUP(J5,$J$8:$K$876,2,FALSE)'

    def generate(self):
        """全シートを生成してワークブックを返す"""
        self.generate_cover_sheet()
        self.generate_overview_sheet()
        self.generate_screen_layout_sheets()
        self.generate_processing_sheet()
        self.generate_revision_history_sheet()
        return self.wb

    def save(self, output_path):
        """ワークブックをExcelファイルに保存"""
        self.wb.save(output_path)
        print(f"✓ Excel ファイルが正常に保存されました: {output_path}")


# ============================================================================
# Helper Functions
# ============================================================================

def generate_from_config(config_dict, output_path):
    """
    設定ディクショナリからExcelファイルを生成

    Args:
        config_dict (dict): 設定ディクショナリ
        output_path (str|Path): 出力ファイルパス

    Returns:
        Path: 生成されたファイルパス
    """
    generator = SpecGenerator(config_dict)
    generator.generate()
    generator.save(output_path)
    return Path(output_path)


def load_config_from_json(json_path):
    """JSON設定ファイルを読み込む"""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


# ============================================================================
# Sample Configuration
# ============================================================================

SAMPLE_CONFIG = {
    "cover": {
        "company": "株式会社エフエム",
        "project": "Socia2026",
        "system": "Sociaポータル",
        "function_name": "(共通)現職エリア",
        "function_id": "gensyokuArea",
        "author": "ISI李",
        "create_date": "2024-11-15",
        "update_date": "2026-01-19",
        "update_author": "ISI孫"
    },
    "overview": {
        "content": "この機能は、一覧表示画面から現職エリアの詳細情報を表示および編集する機能です。"
    },
    "screens": [
        {
            "name": "現職エリア",
            "objects": [
                {
                    "no": 1,
                    "name": "エリアタイトル",
                    "type": "ラベル",
                    "attributes": ["フォント: 14pt", "色: 黒"],
                    "actions": [],
                    "display_controls": [],
                    "error_definitions": []
                }
            ]
        }
    ],
    "processing": {
        "apis": [
            {
                "component_name": "現職エリア（gensyokuarea）",
                "action_type": "初期表示",
                "name": "(共通)現職エリア-表示",
                "id": "GetCurrentAreaInfo",
                "request_params": [
                    {
                        "name": "社員コードリスト",
                        "description": "表示対象社員の社員コード"
                    }
                ],
                "response_description": "レスポンスデータを画面に表示する",
                "display_mode": "■ブロック表示",
                "response_params": [
                    {
                        "screen_item": "項目名",
                        "response_param": "現職エリアのタイトル",
                        "note": "注記がある場合はここに記入"
                    }
                ]
            }
        ]
    }
}


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Web UI設計書Excelジェネレーター',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
例:
  python generate_spec.py config.json
  python generate_spec.py -c config.json -o output.xlsx
  python generate_spec.py --config config.json --output /path/to/output.xlsx
        """
    )
    parser.add_argument('config', nargs='?', default=None,
                        help='JSON設定ファイルのパス')
    parser.add_argument('-c', '--config', dest='config_alt', default=None,
                        help='JSON設定ファイルのパス（--config フラグ版）')
    parser.add_argument('-o', '--output', default=None,
                        help='出力Excelファイルのパス（デフォルト: spec_generated.xlsx）')

    args = parser.parse_args()

    # Determine config file
    config_path = args.config or args.config_alt
    if not config_path:
        print("エラー: 設定ファイルが指定されていません")
        print("使用方法: python generate_spec.py <config.json>")
        sys.exit(1)

    # Determine output file
    output_path = args.output or 'spec_generated.xlsx'

    try:
        # Load config
        config = load_config_from_json(config_path)

        # Generate Excel file
        result_path = generate_from_config(config, output_path)
        print(f"✓ 完了: {result_path}")

    except FileNotFoundError:
        print(f"エラー: 設定ファイルが見つかりません: {config_path}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"エラー: JSON解析エラー: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"エラー: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"エラー: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
