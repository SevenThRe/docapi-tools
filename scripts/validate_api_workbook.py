from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_SHEETS = [
    "表紙",
    "処理概要",
    "APIシーケンス&DFD",
    "リクエストAPIパラメーター",
    "レスポンスAPIパラメーター",
    "処理詳細",
    "改定履歴",
]


def _sheet_text(ws, *, max_rows: int | None = None) -> str:
    parts: list[str] = []
    row_limit = max_rows or ws.max_row
    for row in ws.iter_rows(min_row=1, max_row=row_limit, values_only=True):
        for value in row:
            if value is None:
                continue
            text = str(value).strip()
            if text:
                parts.append(text)
    return "\n".join(parts)


def validate_api_workbook(
    workbook_path: str | Path,
    *,
    api_config: dict[str, Any] | None = None,
    analysis_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    workbook_file = Path(workbook_path).expanduser().resolve()
    wb = load_workbook(workbook_file)
    issues: list[dict[str, str]] = []

    try:
        def add_issue(severity: str, code: str, message: str, *, sheet: str = "workbook") -> None:
            issues.append(
                {
                    "severity": severity,
                    "code": code,
                    "message": message,
                    "sheet": sheet,
                }
            )

        for sheet_name in REQUIRED_SHEETS:
            if sheet_name not in wb.sheetnames:
                add_issue("error", "missing_sheet", f"必須シート `{sheet_name}` がありません。")

        if "リクエストAPIパラメーター" in wb.sheetnames:
            request_text = _sheet_text(wb["リクエストAPIパラメーター"], max_rows=20)
            if "説明" not in request_text or "必須" not in request_text:
                add_issue("warning", "request_layout", "リクエストパラメーターシートの基本レイアウトが不足しています。", sheet="リクエストAPIパラメーター")

        if "レスポンスAPIパラメーター" in wb.sheetnames:
            response_text = _sheet_text(wb["レスポンスAPIパラメーター"], max_rows=20)
            if "説明" not in response_text or "必須" not in response_text:
                add_issue("warning", "response_layout", "レスポンスパラメーターシートの基本レイアウトが不足しています。", sheet="レスポンスAPIパラメーター")

        if "APIシーケンス&DFD" in wb.sheetnames:
            sequence_text = _sheet_text(wb["APIシーケンス&DFD"])
            if analysis_payload:
                discovered = (analysis_payload.get("analysis") or analysis_payload).get("discovered_files", {})
                has_sql_evidence = bool(discovered.get("mybatis_xml"))
                if has_sql_evidence and not any(token in sequence_text for token in ("DB", "テーブル", "社員申請", "申請")):
                    add_issue("warning", "db_table_visibility", "SQL証跡はあるが、シーケンス上でDB/テーブル表現が弱いです。", sheet="APIシーケンス&DFD")

        if "処理詳細" in wb.sheetnames:
            detail_text = _sheet_text(wb["処理詳細"])
            if api_config:
                detail_steps = (((api_config.get("processing_detail") or {}).get("steps")) or [])
                has_sqlish_content = any(
                    isinstance(content, dict) and content.get("type") in {"sql", "sql_structured", "mybatis_sql"}
                    for step in detail_steps if isinstance(step, dict)
                    for content in step.get("content", [])
                )
                if has_sqlish_content and not any(token in detail_text for token in ("SELECT", "UPDATE", "INSERT", "DELETE", "FROM")):
                    add_issue("warning", "sql_rendering", "処理詳細にSQL由来の記述が十分に展開されていません。", sheet="処理詳細")
            if not any("\u3040" <= ch <= "\u9fff" for ch in detail_text):
                add_issue("warning", "japanese_detail", "処理詳細シートに日本語本文が不足しています。", sheet="処理詳細")

        status = "pass"
        if any(issue["severity"] == "error" for issue in issues):
            status = "fail"
        elif issues:
            status = "review"

        return {
            "status": status,
            "workbook_path": str(workbook_file),
            "required_sheets": REQUIRED_SHEETS,
            "issues": issues,
            "sheet_names": wb.sheetnames,
        }
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate API workbook structure and fidelity signals.")
    parser.add_argument("workbook", help="Workbook path")
    parser.add_argument("--api-config", default=None, help="api_config.json path")
    parser.add_argument("--analysis", default=None, help="analysis.json path")
    parser.add_argument("-o", "--output", default=None, help="Write validation report JSON")
    args = parser.parse_args()

    api_config = json.loads(Path(args.api_config).read_text(encoding="utf-8")) if args.api_config else None
    analysis_payload = json.loads(Path(args.analysis).read_text(encoding="utf-8")) if args.analysis else None
    report = validate_api_workbook(args.workbook, api_config=api_config, analysis_payload=analysis_payload)
    rendered = json.dumps(report, ensure_ascii=False, indent=2)
    print(rendered)
    if args.output:
        Path(args.output).write_text(rendered + "\n", encoding="utf-8")
    raise SystemExit(0 if report["status"] != "fail" else 1)


if __name__ == "__main__":
    main()
