#!/usr/bin/env python3
"""
テンプレートベースのWeb UI設計書ジェネレーター v4
==============================================

使用方法:
    python generate_from_template.py config.json
    python generate_from_template.py config.json -o output.xlsx

アーキテクチャ:
  1. template_clean.xlsx（画像なし）を読み込む
  2. Config検証（config_schema.py）
  3. 表紙シートの値を設定
  4. 外部-機能概要: 機能概要・画面一覧・処理フローを動的行数で生成
  5. 外部-画面レイアウト: オブジェクト定義を生成（種類/属性/動作/表示制御/エラー）
  6. 内部-処理詳細: API定義を生成（50行分页境界線付き）
  7. 外部-画面レイアウト: シート名をconfig変数で差し替え
"""

import json
import sys
import argparse
from pathlib import Path
from datetime import datetime
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

# Config 検証（同ディレクトリの config_schema.py）
try:
    from config_schema import validate_config, validate_and_report
except ImportError:
    sys.path.insert(0, str(Path(__file__).parent))
    try:
        from config_schema import validate_config, validate_and_report
    except ImportError:
        def validate_config(c):
            return [], []
        def validate_and_report(c):
            return True


# ============================================================================
# プロトタイプ行マッピング（template_clean.xlsx の行番号）
# ============================================================================
PROTO = {
    'spacer':        8,   # 空行
    'component':     9,   # B=コンポーネント名
    'spacer2':      10,   # 空行（コンポーネント名の後）
    'action_first': 11,   # C=初期表示（最初のアクション行）
    'api_label':    12,   # D='API名'
    'api_value':    13,   # E=API名の値
    'req_label':    14,   # D='APIリクエスト'
    'req_header':   15,   # E:Q='リクエストパラメーター', R:AT='値'
    'req_data':     16,   # E:Q=パラメーター名, R:AT=説明
    'spacer_api':   19,   # APIリクエスト〜レスポンス間の空行
    'resp_label':   20,   # D='APIレスポンス'
    'resp_desc':    21,   # E=レスポンス説明文
    'resp_mode':    22,   # E='■ブロック表示' 等
    'resp_header':  23,   # E:Q='画面項目', R:AT='レスポンスパラメーター'
    'resp_data':    24,   # E:Q=画面項目, R:AT=レスポンスパラメーター
    'note':         29,   # E=※注記テキスト
    'end_spacer1':  39,   # 末尾空行1
    'end_spacer2':  40,   # 末尾空行2
    'sub_action':   41,   # B=番号, C=アクション名（2番目以降のアクション）
}

# マージパターン（E:Q + R:AT の同一行マージ）
MERGE_EQ_RAT = [('E', 'Q'), ('R', 'AT')]

# 分页境界線の間隔（行数）と最初の境界行
PAGE_SEP_FIRST = 51    # 最初の分页境界行（ヘッダー7行 + コンテンツ44行）
PAGE_SEP_INTERVAL = 50 # 以降50行ごと

# 分页境界線の適用範囲（列A〜AZ）
PAGE_SEP_LAST_COL = column_index_from_string('AZ')

PROCESSING_FONT_NAME = "Meiryo"
PROCESSING_BODY_SIZE = 10
PROCESSING_HEADER_SIZE = 10.5

# ============================================================================
# 画面レイアウトシートのプロトタイプ行マッピング
# テンプレートの「外部-画面レイアウト-現職エリア」から取得
# ============================================================================
LAYOUT_PROTO_SHEET = "外部-画面レイアウト-現職エリア"

# 画面レイアウト用プロトタイプ行（テンプレート内の行番号）
LAYOUT_PROTO = {
    'screen_id':     95,   # C=画面ID, H=値
    'screen_name':   96,   # C=画面名, H=値
    'spacer':        97,   # 空行
    'init_header':   98,   # C=初期表示
    'def_dh':       104,   # D=定義ラベル, H=値（種類/属性/動作/表示制御/API）
    'cont_h':       107,   # H=値のみ（動作の継続行）
    'obj_header':   103,   # B=No, C=名前（オブジェクトヘッダー）
    'sub_def_eh':   119,   # E=サブ定義ラベル, H=値（ネスト構造）
    'err_header':   127,   # E=エラー, H:AF=条件見出し, AG:AQ=メッセージID, AR:AW=表示ｵﾌﾞｼﾞｪｸﾄNo
    'err_data':     128,   # H:AF=条件, AG:AQ=MSG_XXXXX, AR:AW=No
}

# エラーテーブルのマージパターン
MERGE_ERR = [('H', 'AF'), ('AG', 'AQ'), ('AR', 'AW')]


# ============================================================================
# ヘルパー関数
# ============================================================================

def format_date(value):
    """日付を date オブジェクトに変換"""
    if isinstance(value, str):
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except Exception:
            return value
    if isinstance(value, datetime):
        return value.date()
    return value


def copy_cell_style(src, tgt):
    """セルスタイルをコピー（値はコピーしない）"""
    if src.has_style:
        tgt.font = copy(src.font)
        tgt.border = copy(src.border)
        tgt.fill = copy(src.fill)
        tgt.number_format = src.number_format
        tgt.alignment = copy(src.alignment)
    else:
        tgt.font = Font()
        tgt.border = Border()
        tgt.fill = PatternFill()
        tgt.alignment = Alignment()


def clear_row(ws, row_num, max_col):
    """指定行の値とスタイルをクリア"""
    for col in range(1, max_col + 1):
        try:
            cell = ws.cell(row=row_num, column=col)
            cell.value = None
            cell.font = Font()
            cell.border = Border()
            cell.fill = PatternFill()
            cell.alignment = Alignment()
        except AttributeError:
            pass


def apply_page_separator(ws, row_num, last_col=PAGE_SEP_LAST_COL):
    """
    指定行の全列（A〜AZ）にthin底辺ボーダーを追加して分页境界線を引く。
    既存のボーダーを保持しつつ bottom=thin を上書きする。
    """
    thin = Side(style='thin')
    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        existing = cell.border
        # 既存の左右上ボーダーを保持して bottom だけ追加
        new_border = Border(
            left=existing.left if existing and existing.left else Side(),
            right=existing.right if existing and existing.right else Side(),
            top=existing.top if existing and existing.top else Side(),
            bottom=thin,
            diagonal=existing.diagonal if existing else None,
            diagonal_direction=existing.diagonal_direction if existing else None,
        )
        cell.border = new_border


# ============================================================================
# メインクラス
# ============================================================================

class TemplateGenerator:
    """テンプレートベースのUI設計書ジェネレーター"""

    def __init__(self, config, template_path=None):
        self.config = config

        if template_path is None:
            template_path = Path(__file__).parent.parent / "assets" / "template_clean.xlsx"

        self.tpl_path = Path(template_path)
        if not self.tpl_path.exists():
            raise FileNotFoundError(f"テンプレートが見つかりません: {template_path}")

        self.wb = load_workbook(self.tpl_path)
        self._proto_wb = load_workbook(self.tpl_path)
        self._proto_ws = self._proto_wb["内部-処理詳細"]
        self._proto_max_col = self._proto_ws.max_column

        # 画面レイアウト用プロトタイプ
        if LAYOUT_PROTO_SHEET in self._proto_wb.sheetnames:
            self._layout_proto_ws = self._proto_wb[LAYOUT_PROTO_SHEET]
            self._layout_proto_max_col = self._layout_proto_ws.max_column
        else:
            self._layout_proto_ws = None
            self._layout_proto_max_col = 54  # AZ=52 だが余裕を持つ

        self._strip_images()

    def _strip_images(self):
        """全シートから画像を除去"""
        for sname in self.wb.sheetnames:
            ws = self.wb[sname]
            if hasattr(ws, '_images'):
                ws._images = []

    # ----------------------------------------------------------------
    # プロトタイプ行操作
    # ----------------------------------------------------------------

    def _copy_proto_row(self, ws, target_row, proto_key, values=None, merges=None):
        """
        テンプレートのプロトタイプ行スタイルをtarget_rowにコピーし、
        valuesで指定されたセルに値を設定する。

        Args:
            ws: 書き込み先ワークシート
            target_row: 書き込み先行番号
            proto_key: PROTO辞書のキー（または直接行番号）
            values: {列文字: 値} の辞書
            merges: [(開始列文字, 終了列文字), ...] 同一行のマージパターン
        """
        proto_row = PROTO[proto_key] if isinstance(proto_key, str) else proto_key

        for col_idx in range(1, self._proto_max_col + 1):
            src = self._proto_ws.cell(row=proto_row, column=col_idx)
            tgt = ws.cell(row=target_row, column=col_idx)
            copy_cell_style(src, tgt)
            tgt.value = None

        # 行高さコピー（Noneの場合はデフォルト高さにリセット）
        h = self._proto_ws.row_dimensions[proto_row].height
        ws.row_dimensions[target_row].height = h

        if merges:
            for (sc, ec) in merges:
                try:
                    ws.merge_cells(f"{sc}{target_row}:{ec}{target_row}")
                except Exception:
                    pass

        if values:
            for col_letter, val in values.items():
                try:
                    ws[f"{col_letter}{target_row}"] = val
                except AttributeError:
                    pass

    def _set_processing_text_style(
        self,
        ws,
        row,
        columns,
        *,
        size=None,
        bold=None,
        horizontal=None,
        vertical="center",
        wrap_text=None,
    ):
        for col in columns:
            cell = ws[f"{col}{row}"]
            font = copy(cell.font)
            font.name = PROCESSING_FONT_NAME
            if size is not None:
                font.sz = size
            if bold is not None:
                font.bold = bold
            cell.font = font

            alignment = copy(cell.alignment)
            if horizontal is not None:
                alignment.horizontal = horizontal
            if vertical is not None:
                alignment.vertical = vertical
            if wrap_text is not None:
                alignment.wrap_text = wrap_text
            cell.alignment = alignment

    def _format_processing_row(self, ws, row, kind):
        return

    def _polish_processing_sheet(self, ws):
        return

    # ----------------------------------------------------------------
    # 表紙シート
    # ----------------------------------------------------------------

    def fill_cover(self):
        """表紙シートに設定値を書き込む"""
        ws = self.wb["表紙"]
        cover = self.config.get("cover", {})

        mapping = {
            "C4":  cover.get("company", ""),
            "C5":  cover.get("project", ""),
            "C6":  cover.get("system", ""),
            "C7":  cover.get("function_name", ""),
            "C8":  cover.get("function_id", ""),
            "C10": format_date(cover.get("create_date", "")),
            "C11": cover.get("author", "${username}"),
            "C12": format_date(cover.get("update_date", "")),
            "C13": cover.get("update_author", "${username}"),
        }
        for cell_ref, val in mapping.items():
            ws[cell_ref] = val

    # ----------------------------------------------------------------
    # 外部-機能概要シート
    # ----------------------------------------------------------------

    def fill_overview(self):
        """
        外部-機能概要シートに内容を書き込む。
        動的行数対応: screens が何件あっても正しく配置する。

        レイアウト:
          B8  = '機能概要'（固定ラベル）
          H9  = 機能概要説明文
          B11 = '画面一覧'（固定ラベル）
          C12 = 1件目のコンポーネント名, H13 = 説明
          C{n}= 2件目のコンポーネント名, H{n+1} = 説明  ← 動的に追加
          ...
          B{m}  = '処理フロー'（ラベル。screens数で位置が変動）
          H{m+1} = 処理フロー説明文
        """
        if "外部-機能概要" not in self.wb.sheetnames:
            return
        ws = self.wb["外部-機能概要"]
        overview = self.config.get("overview", {})

        # --- 行8以降をクリア（テンプレート既存データの残留を防止） ---
        for mc in list(ws.merged_cells.ranges):
            if mc.min_row >= 8:
                ws.unmerge_cells(str(mc))

        max_col = ws.max_column
        for row in range(8, max(ws.max_row + 1, 30)):  # 少なくとも30行目までクリア
            clear_row(ws, row, max_col)

        # --- 固定ラベルを再配置 ---
        ws["B8"] = '機能概要'
        ws["B8"].font = Font(name='Kosugi', size=9)

        # H9: 機能概要説明文
        description = overview.get("description", "")
        if description:
            ws["H9"] = description
            ws["H9"].alignment = Alignment(horizontal="left", vertical="top",
                                           wrap_text=True)
            ws["H9"].font = Font(name='Kosugi', size=9)

        # B11: 画面一覧ラベル
        ws["B11"] = '画面一覧'
        ws["B11"].font = Font(name='Kosugi', size=9)

        # 画面一覧（C列=コンポーネント名, H列=説明）
        screens = overview.get("screens", [])
        cur = 12  # 最初のコンポーネント行
        for screen in screens:
            screen_name = screen.get("name", "")
            descriptions = screen.get("descriptions", [])

            # コンポーネント名
            ws.cell(row=cur, column=3).value = screen_name  # C列
            ws.cell(row=cur, column=3).font = Font(name='Kosugi', size=9)
            cur += 1

            # 各説明文
            for desc in descriptions:
                ws.cell(row=cur, column=8).value = desc  # H列
                ws.cell(row=cur, column=8).font = Font(name='Kosugi', size=9)
                ws.cell(row=cur, column=8).alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True)
                cur += 1

            # コンポーネント間スペーサー
            cur += 1

        # 処理フロー（screens数で位置が変動するので動的に配置）
        flow_row_label = max(cur, 18)  # 最低でも行18以降
        ws.cell(row=flow_row_label, column=2).value = '処理フロー'
        ws.cell(row=flow_row_label, column=2).font = Font(name='Kosugi', size=9)

        flow_desc = overview.get("flow_description", "")
        if flow_desc:
            ws.cell(row=flow_row_label + 1, column=8).value = flow_desc
            ws.cell(row=flow_row_label + 1, column=8).alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True)
            ws.cell(row=flow_row_label + 1, column=8).font = Font(name='Kosugi', size=9)

    # ----------------------------------------------------------------
    # 画面レイアウトシート名の変数差し替え
    # ----------------------------------------------------------------

    def rename_layout_sheets(self):
        """
        外部-画面レイアウトシートの名前をconfigの変数で差し替える。

        Config の screen_layouts セクション:
          [
            {
              "template_sheet": "外部-画面レイアウト-現職エリア",   # テンプレート内の既存シート名
              "output_name": "外部-画面レイアウト-社員検索",         # 出力時のシート名
              "title": "社員検索画面"                               # 任意の表示タイトル
            }
          ]

        シート名命名規則:
          - 機能レベルの画面: "外部-画面レイアウト-{機能名称}"
            例) "外部-画面レイアウト-社員検索"
          - UIコンポーネントレベル: "外部-画面レイアウト-{UI部品名称}"
            例) "外部-画面レイアウト-現職エリア", "外部-画面レイアウト-経歴モーダル"
        """
        screen_layouts = self.config.get("screen_layouts", [])
        for layout in screen_layouts:
            tpl_name = layout.get("template_sheet", "")
            out_name = layout.get("output_name", "")
            if tpl_name and out_name and tpl_name in self.wb.sheetnames:
                ws = self.wb[tpl_name]
                ws.title = out_name

    # ----------------------------------------------------------------
    # 画面レイアウト用プロトタイプ行コピー
    # ----------------------------------------------------------------

    def _copy_layout_proto_row(self, ws, target_row, proto_key, values=None, merges=None):
        """画面レイアウトシートのプロトタイプ行スタイルをコピーする"""
        if self._layout_proto_ws is None:
            # フォールバック: 基本スタイルで書き込み
            if values:
                for col_letter, val in values.items():
                    cell = ws[f"{col_letter}{target_row}"]
                    cell.value = val
                    cell.font = Font(name='Kosugi', size=9)
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
            return

        proto_row = LAYOUT_PROTO[proto_key] if isinstance(proto_key, str) else proto_key

        for col_idx in range(1, self._layout_proto_max_col + 1):
            src = self._layout_proto_ws.cell(row=proto_row, column=col_idx)
            tgt = ws.cell(row=target_row, column=col_idx)
            copy_cell_style(src, tgt)
            tgt.value = None

        h = self._layout_proto_ws.row_dimensions[proto_row].height
        ws.row_dimensions[target_row].height = h

        if merges:
            for (sc, ec) in merges:
                try:
                    ws.merge_cells(f"{sc}{target_row}:{ec}{target_row}")
                except Exception:
                    pass

        if values:
            for col_letter, val in values.items():
                try:
                    ws[f"{col_letter}{target_row}"] = val
                except AttributeError:
                    pass

    # ----------------------------------------------------------------
    # 画面レイアウトシートの生成
    # ----------------------------------------------------------------

    def generate_screen_layouts(self):
        """
        各画面のオブジェクト定義を画面レイアウトシートに書き込む。

        Config の screens セクション:
          [
            {
              "name": "現職エリア",
              "target_sheet": "外部-画面レイアウト-現職エリア",  # 書き込み先シート名
              "screen_id": "gensyokuarea",                      # 画面ID
              "screen_name": "(共通)現職エリア",                  # 画面名
              "initial_display": {                               # 初期表示セクション（省略可）
                "actions": ["ブロックモードまたは一覧モードで表示する"],
                "api": ["表示API(id：gensyokuArea/showBlock)を呼出し"]
              },
              "objects": [
                {
                  "no": 1,
                  "name": "社員切替",
                  "type": "アイコン",
                  "attributes": ["cursor：pointer"],
                  "actions": [
                    "上アイコンを押下する時、直前社員へ切替",
                    "先頭の社員の場合、末尾の社員へ切り替える"
                  ],
                  "display_controls": ["一覧モードで複数人の場合は非活性とする"],
                  "api": "現職エリアの表示APIを呼出し",
                  "sub_items": [                                 # ネスト要素（省略可）
                    {
                      "label": "切替範囲",
                      "definitions": [
                        {"item": "種類", "value": "テキスト"},
                        {"item": "動作", "value": "..."}
                      ]
                    }
                  ],
                  "error_definitions": [
                    {
                      "header": "押下時のエラーチェック",
                      "condition": "該当する社員がいない場合",
                      "message_id": "MSG_00089",
                      "object_no": "1"
                    }
                  ]
                }
              ]
            }
          ]
        """
        screens = self.config.get("screens", [])
        if not screens:
            return

        for screen in screens:
            target_sheet = screen.get("target_sheet", "")
            if not target_sheet or target_sheet not in self.wb.sheetnames:
                print(f"  [warn] skip: sheet '{target_sheet}' was not found")
                continue

            ws = self.wb[target_sheet]
            screen_id = screen.get("screen_id", "")
            screen_name = screen.get("screen_name", "")
            objects = screen.get("objects", [])
            initial_display = screen.get("initial_display", {})

            # 行8以降のマージセルを解除
            for mc in list(ws.merged_cells.ranges):
                if mc.min_row >= 8:
                    ws.unmerge_cells(str(mc))

            # 行8以降をクリア
            max_col = ws.max_column
            for row in range(8, ws.max_row + 1):
                clear_row(ws, row, max_col)

            # --- 書き込み開始 ---
            cur = 8

            # 画面ID・画面名
            self._copy_layout_proto_row(ws, cur, 'screen_id',
                                        values={'C': '画面ID', 'H': screen_id})
            cur += 1
            self._copy_layout_proto_row(ws, cur, 'screen_name',
                                        values={'C': '画面名', 'H': screen_name})
            cur += 1

            # 空行
            self._copy_layout_proto_row(ws, cur, 'spacer')
            cur += 1

            # 初期表示セクション
            if initial_display:
                self._copy_layout_proto_row(ws, cur, 'init_header',
                                            values={'C': '初期表示'})
                cur += 1

                # 動作
                init_actions = initial_display.get("actions", [])
                if init_actions:
                    self._copy_layout_proto_row(ws, cur, 'def_dh',
                                                values={'D': '動作', 'H': init_actions[0]})
                    cur += 1
                    for action in init_actions[1:]:
                        self._copy_layout_proto_row(ws, cur, 'cont_h',
                                                    values={'H': action})
                        cur += 1

                # API
                init_apis = initial_display.get("api", [])
                if init_apis:
                    self._copy_layout_proto_row(ws, cur, 'def_dh',
                                                values={'D': 'API', 'H': init_apis[0]})
                    cur += 1
                    for api_line in init_apis[1:]:
                        self._copy_layout_proto_row(ws, cur, 'cont_h',
                                                    values={'H': api_line})
                        cur += 1

                # 空行
                self._copy_layout_proto_row(ws, cur, 'spacer')
                cur += 1

            # --- オブジェクト定義 ---
            for obj in objects:
                obj_no = obj.get("no", "")
                obj_name = obj.get("name", "")

                # オブジェクトヘッダー行
                self._copy_layout_proto_row(ws, cur, 'obj_header',
                                            values={'B': obj_no, 'C': obj_name})
                cur += 1

                # 種類
                obj_type = obj.get("type", "")
                if obj_type:
                    self._copy_layout_proto_row(ws, cur, 'def_dh',
                                                values={'D': '種類', 'H': obj_type})
                    cur += 1

                # 属性
                attributes = obj.get("attributes", [])
                if attributes:
                    self._copy_layout_proto_row(ws, cur, 'def_dh',
                                                values={'D': '属性', 'H': attributes[0]})
                    cur += 1
                    for attr in attributes[1:]:
                        self._copy_layout_proto_row(ws, cur, 'cont_h',
                                                    values={'H': attr})
                        cur += 1

                # 動作
                actions = obj.get("actions", [])
                if actions:
                    self._copy_layout_proto_row(ws, cur, 'def_dh',
                                                values={'D': '動作', 'H': actions[0]})
                    cur += 1
                    for action in actions[1:]:
                        self._copy_layout_proto_row(ws, cur, 'cont_h',
                                                    values={'H': action})
                        cur += 1

                # 表示制御
                display_controls = obj.get("display_controls", [])
                if display_controls:
                    self._copy_layout_proto_row(ws, cur, 'def_dh',
                                                values={'D': '表示制御', 'H': display_controls[0]})
                    cur += 1
                    for ctrl in display_controls[1:]:
                        self._copy_layout_proto_row(ws, cur, 'cont_h',
                                                    values={'H': ctrl})
                        cur += 1

                # API
                api_ref = obj.get("api", "")
                if api_ref:
                    if isinstance(api_ref, list):
                        self._copy_layout_proto_row(ws, cur, 'def_dh',
                                                    values={'D': 'API', 'H': api_ref[0]})
                        cur += 1
                        for api_line in api_ref[1:]:
                            self._copy_layout_proto_row(ws, cur, 'cont_h',
                                                        values={'H': api_line})
                            cur += 1
                    else:
                        self._copy_layout_proto_row(ws, cur, 'def_dh',
                                                    values={'D': 'API', 'H': api_ref})
                        cur += 1

                # サブ項目（ネスト構造 E列）
                sub_items = obj.get("sub_items", [])
                for sub in sub_items:
                    sub_label = sub.get("label", "")
                    self._copy_layout_proto_row(ws, cur, 'spacer')
                    cur += 1

                    # サブ項目ヘッダー（D列にラベル）
                    self._copy_layout_proto_row(ws, cur, 'def_dh',
                                                values={'D': sub_label})
                    cur += 1

                    # サブ定義（E列にラベル、H列に値）
                    for defn in sub.get("definitions", []):
                        self._copy_layout_proto_row(ws, cur, 'sub_def_eh',
                                                    values={
                                                        'E': defn.get("item", ""),
                                                        'H': defn.get("value", "")
                                                    })
                        cur += 1

                    # サブエラー定義
                    for err in sub.get("error_definitions", []):
                        self._write_error_row(ws, cur, err)
                        cur += 2  # ヘッダー行 + データ行

                    # サブ表示制御
                    for ctrl_text in sub.get("display_controls", []):
                        self._copy_layout_proto_row(ws, cur, 'sub_def_eh',
                                                    values={'E': '表示制御', 'H': ctrl_text})
                        cur += 1

                    # サブAPI
                    sub_api = sub.get("api", "")
                    if sub_api:
                        self._copy_layout_proto_row(ws, cur, 'sub_def_eh',
                                                    values={'E': 'API', 'H': sub_api})
                        cur += 1

                # エラー定義（トップレベル）
                error_defs = obj.get("error_definitions", [])
                if error_defs:
                    for err in error_defs:
                        self._write_error_row(ws, cur, err)
                        cur += 2  # ヘッダー行 + データ行

                # オブジェクト間スペーサー
                self._copy_layout_proto_row(ws, cur, 'spacer')
                cur += 1

            # 分页境界線を適用
            total_rows = cur
            sep_row = PAGE_SEP_FIRST
            while sep_row <= total_rows:
                apply_page_separator(ws, sep_row)
                sep_row += PAGE_SEP_INTERVAL

            print(f"  [ok] {target_sheet}: {len(objects)} objects, {cur - 8} generated rows")

    def _write_error_row(self, ws, cur_row, err_def):
        """エラー定義テーブル（ヘッダー行 + データ行）を書き込む"""
        header_text = err_def.get("header", "エラーチェック")
        condition = err_def.get("condition", "")
        message_id = err_def.get("message_id", "")
        object_no = err_def.get("object_no", "")

        # ヘッダー行
        self._copy_layout_proto_row(
            ws, cur_row, 'err_header',
            values={'E': 'エラー', 'H': header_text, 'AG': 'メッセージID', 'AR': '表示ｵﾌﾞｼﾞｪｸﾄNo'},
            merges=MERGE_ERR
        )
        # データ行
        self._copy_layout_proto_row(
            ws, cur_row + 1, 'err_data',
            values={'H': condition, 'AG': message_id, 'AR': object_no},
            merges=MERGE_ERR
        )

    # ----------------------------------------------------------------
    # 内部-処理詳細シート
    # ----------------------------------------------------------------

    def generate_processing(self):
        """
        内部-処理詳細シートを生成する。

        Config の processing セクション:
          {
            "component_name": "現職エリア（gensyokuarea）",  # B9に表示
            "apis": [
              {
                "name": "API名",
                "action_type": "初期表示",        # 省略時は "初期表示"
                "action_number": null,            # 2番目以降はB列の番号（1,2...）
                "request_params": [
                  {"name": "パラメーター名", "description": "説明"}
                ],
                "response_description": "レスポンス全体の説明文",
                "display_mode": "■ブロック表示",  # 省略時は "■ブロック表示"
                "response_params": [
                  {
                    "screen_item": "画面項目名",
                    "response_param": "レスポンスパラメーター名 or 説明",
                    "note": "AA列の条件注記（省略可）"
                  }
                ],
                "notes": [
                  "※レスポンスパラメーター「xxx」がnoneの場合...",
                  "※対象社員の顔写真がない場合..."
                ]
              }
            ]
          }

        ※notes の生成ガイドライン（AI向けhook）:
          - response_params に条件分岐・例外動作がある場合は必ず notes に追記
          - 「※{パラメーター名}が{値}の場合、{動作}」の書式を使う
          - 複数条件は行を分けて記述する
        """
        if "内部-処理詳細" not in self.wb.sheetnames:
            return

        ws = self.wb["内部-処理詳細"]
        processing = self.config.get("processing", {})
        apis = processing.get("apis", [])
        component_name = processing.get("component_name", "")

        if not apis and not component_name:
            return

        # 行8以降のマージセルを解除
        for mc in list(ws.merged_cells.ranges):
            if mc.min_row >= 8:
                ws.unmerge_cells(str(mc))

        # 行8以降をクリア
        max_col = ws.max_column
        for row in range(8, ws.max_row + 1):
            clear_row(ws, row, max_col)

        # --- 書き込み開始 ---
        cur = 8

        self._copy_proto_row(ws, cur, 'spacer')
        self._format_processing_row(ws, cur, "spacer")
        cur += 1

        self._copy_proto_row(ws, cur, 'component', values={'B': component_name})
        self._format_processing_row(ws, cur, "component")
        cur += 1

        self._copy_proto_row(ws, cur, 'spacer2')
        self._format_processing_row(ws, cur, "spacer")
        cur += 1

        for api_idx, api in enumerate(apis):
            action_type   = api.get("action_type", "初期表示")
            action_number = api.get("action_number", None)

            # アクション行
            if api_idx == 0:
                self._copy_proto_row(ws, cur, 'action_first',
                                     values={'C': action_type})
            else:
                self._copy_proto_row(ws, cur, 'sub_action',
                                     values={
                                         'B': action_number if action_number is not None else api_idx,
                                         'C': action_type
                                     })
            self._format_processing_row(ws, cur, "action")
            cur += 1

            # API名
            self._copy_proto_row(ws, cur, 'api_label', values={'D': 'API名'})
            self._format_processing_row(ws, cur, "section_label")
            cur += 1
            self._copy_proto_row(ws, cur, 'api_value',
                                 values={'E': api.get("name", "")})
            self._format_processing_row(ws, cur, "api_value")
            cur += 1

            # APIリクエスト
            self._copy_proto_row(ws, cur, 'req_label', values={'D': 'APIリクエスト'})
            self._format_processing_row(ws, cur, "section_label")
            cur += 1
            self._copy_proto_row(ws, cur, 'req_header',
                                 values={'E': 'リクエストパラメーター', 'R': '値'},
                                 merges=MERGE_EQ_RAT)
            self._format_processing_row(ws, cur, "header")
            cur += 1

            for param in api.get("request_params", []):
                self._copy_proto_row(ws, cur, 'req_data',
                                     values={
                                         'E': param.get("name", ""),
                                         'R': param.get("description", "")
                                     },
                                     merges=MERGE_EQ_RAT)
                self._format_processing_row(ws, cur, "request_data")
                cur += 1

            # スペーサー
            self._copy_proto_row(ws, cur, 'spacer_api')
            self._format_processing_row(ws, cur, "spacer")
            cur += 1

            # APIレスポンス
            self._copy_proto_row(ws, cur, 'resp_label', values={'D': 'APIレスポンス'})
            self._format_processing_row(ws, cur, "section_label")
            cur += 1
            self._copy_proto_row(ws, cur, 'resp_desc',
                                 values={'E': api.get("response_description", "")})
            self._format_processing_row(ws, cur, "response_desc")
            cur += 1
            self._copy_proto_row(ws, cur, 'resp_mode',
                                 values={'E': api.get("display_mode", "■ブロック表示")})
            self._format_processing_row(ws, cur, "response_mode")
            cur += 1

            # レスポンスパラメーターヘッダー
            self._copy_proto_row(ws, cur, 'resp_header',
                                 values={'E': '画面項目', 'R': 'レスポンスパラメーター'},
                                 merges=MERGE_EQ_RAT)
            self._format_processing_row(ws, cur, "header")
            cur += 1

            # レスポンスデータ行
            for rparam in api.get("response_params", []):
                values = {
                    'E': rparam.get("screen_item", ""),
                    'R': rparam.get("response_param", "")
                }
                note_inline = rparam.get("note", "")
                if note_inline:
                    # note が既に ※ で始まっていない場合は付与
                    values['AA'] = note_inline if note_inline.startswith('※') else f"※{note_inline}"

                self._copy_proto_row(ws, cur, 'resp_data',
                                     values=values,
                                     merges=MERGE_EQ_RAT)
                self._format_processing_row(ws, cur, "response_data")
                cur += 1

            # ※ 注記行（APIブロック末尾）
            for note_text in api.get("notes", []):
                # ※ プレフィックスを確保
                if note_text and not note_text.startswith('※'):
                    note_text = f"※{note_text}"
                self._copy_proto_row(ws, cur, 'note', values={'E': note_text})
                self._format_processing_row(ws, cur, "note")
                cur += 1

            # APIブロック間スペーサー
            self._copy_proto_row(ws, cur, 'end_spacer1')
            self._format_processing_row(ws, cur, "spacer")
            cur += 1

        # 末尾スペーサー
        self._copy_proto_row(ws, cur, 'end_spacer2')
        self._format_processing_row(ws, cur, "spacer")
        cur += 1

        # ----------------------------------------------------------------
        # 分页境界線を追加（50行ごと: 行51, 101, 151 ...）
        # ----------------------------------------------------------------
        total_rows = cur
        sep_row = PAGE_SEP_FIRST
        while sep_row <= total_rows:
            apply_page_separator(ws, sep_row)
            sep_row += PAGE_SEP_INTERVAL

        # 最終行にも境界線
        if total_rows > 0 and total_rows != sep_row - PAGE_SEP_INTERVAL:
            apply_page_separator(ws, total_rows)

        self._polish_processing_sheet(ws)

    # ----------------------------------------------------------------
    # 保存
    # ----------------------------------------------------------------

    def save(self, output_path):
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(str(out))
        print(f"[ok] excel saved: {out}")

    def generate(self, output_path):
        """全シートを生成してファイルを保存する"""
        # Config 検証
        errors, warnings = validate_config(self.config)
        if warnings:
            for w in warnings:
                print(f"  [warn] {w}")
        if errors:
            print(f"  [error] config validation errors: {len(errors)}")
            for e in errors:
                print(f"  [error] {e}")
            print("  → エラーを修正するか、--force で強制生成してください")
            # エラーがあっても生成は続行する（最低限の出力を得るため）

        self.fill_cover()
        self.fill_overview()
        self.rename_layout_sheets()      # シート名確定を先に行う
        self.generate_screen_layouts()   # リネーム後のシート名で書き込む
        self.generate_processing()
        self.save(output_path)
        return str(output_path)


# ============================================================================
# エントリーポイント
# ============================================================================

def merge_project_config(config, project_config):
    """Merge project-level config into screen config.

    Project config provides defaults for cover fields.
    Screen config values always take precedence.
    """
    if not project_config:
        return config

    cover = config.setdefault('cover', {})
    company = project_config.get('company', {})
    project = project_config.get('project', {})
    author = project_config.get('author', {})

    if not cover.get('company') and company.get('name'):
        cover['company'] = company['name']
    if not cover.get('project') and project.get('name'):
        cover['project'] = project['name']
    if not cover.get('system') and project.get('system_name'):
        cover['system'] = project['system_name']
    if not cover.get('author') and author.get('name'):
        cover['author'] = author['name']

    from datetime import datetime
    today = datetime.now().strftime('%Y-%m-%d')
    if not cover.get('create_date'):
        cover['create_date'] = today
    if not cover.get('update_date'):
        cover['update_date'] = today
    if not cover.get('update_author'):
        cover['update_author'] = cover.get('author', '')

    return config


def main():
    parser = argparse.ArgumentParser(
        description="テンプレートベースのWeb UI設計書ジェネレーター"
    )
    parser.add_argument("config", help="設定JSONファイルのパス")
    parser.add_argument("-p", "--project-config", default=None,
                        help="プロジェクト共通設定JSONファイルのパス")
    parser.add_argument("-o", "--output", default="output.xlsx",
                        help="出力ファイルパス (デフォルト: output.xlsx)")
    parser.add_argument("-t", "--template", default=None,
                        help="テンプレートファイルのパス（省略時は自動検索）")
    args = parser.parse_args()

    config_path = Path(args.config)
    if not config_path.exists():
        print(f"エラー: 設定ファイルが見つかりません: {args.config}", file=sys.stderr)
        sys.exit(1)

    with open(config_path, encoding="utf-8") as f:
        config = json.load(f)

    # Load and merge project config
    pc_path = args.project_config
    if not pc_path:
        auto_path = Path(__file__).parent.parent / 'configs' / 'project_config.json'
        if auto_path.exists():
            pc_path = str(auto_path)

    if pc_path:
        import json as json_mod
        project_config = json_mod.loads(Path(pc_path).read_text(encoding='utf-8'))
        config = merge_project_config(config, project_config)
        print(f'  [ok] project config loaded: {pc_path}')

    template_path = args.template
    if not template_path:
        if pc_path:
            pc = json.loads(Path(pc_path).read_text(encoding='utf-8'))
            tp = pc.get('template_paths', {}).get('ui_spec')
            if tp:
                template_path = str(Path(__file__).parent.parent / tp)

    gen = TemplateGenerator(config, template_path=template_path)
    gen.generate(args.output)
    print(f"[ok] completed: {args.output}")


if __name__ == "__main__":
    main()
