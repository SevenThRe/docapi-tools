#!/usr/bin/env python3
"""
API設計書 Excel ジェネレーター v2
===============================
1つのAPIエンドポイント単位で、日本のSIer標準フォーマットの
API設計書（Excel）を生成する。

v2 改善点:
  - 処理詳細: wrap_text=False でテキスト溢れ表示（原本準拠）
  - パラメーター: wrap_text=False, vertical=center（原本準拠）
  - 処理概要: Excel図形（DrawingML XML注入）で編集可能なフローチャート
  - APIシーケンス&DFD: 矢印コネクター追加（DrawingML）

シート構成:
  - 表紙
  - 処理概要          （テキスト＋Excelフローチャート図形）
  - APIシーケンス&DFD （シーケンステキスト＋矢印コネクター）
  - リクエストAPIパラメーター（階層インデント付き）
  - レスポンスAPIパラメーター（同上）
  - 処理詳細          （疑似コード＋SQL、テキスト溢れ表示）
  - 改定履歴

Usage:
    python generate_api_spec.py config.json [-t template.xlsx] [-o output.xlsx]
"""

import argparse
import json
import io
import os
import re
import sys
import shutil
import tempfile
import zipfile
from pathlib import Path
from copy import copy
from datetime import datetime
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

FONT_HEADER = Font(name='Kosugi', size=9, bold=True)
FONT_NORMAL = Font(name='Kosugi', size=9)
FONT_META   = Font(name='Kosugi', size=8)    # header rows 2-5 metadata only
FONT_TITLE  = Font(name='Kosugi', size=20)   # sheet title (row 1)

# ★ v2: wrap_text=False (原本準拠 - テキストはセルを溢れて表示)
ALIGN_VCL   = Alignment(vertical='center', horizontal='left', wrap_text=False)
ALIGN_VCC   = Alignment(vertical='center', horizontal='center', wrap_text=False)
ALIGN_VTL   = Alignment(vertical='top', horizontal='left', wrap_text=False)

JAPANESE_CHAR_RE = re.compile(r'[\u3040-\u30ff\u31f0-\u31ff\u3400-\u4dbf\u4e00-\u9fff\u3005\u30fc]')
# Common Simplified-Chinese phrases that should not appear in Japanese-facing prose.
DISALLOWED_SIMPLIFIED_PHRASES = (
    '代码',
    '数据',
    '对象',
    '页面',
    '请求参数',
    '响应参数',
)

# 処理詳細: indent column mapping
# depth 0 → col B(2), depth 1 → col C(3), ..., depth 6 → col H(8)
DETAIL_BASE_COL = 2
# Arrow offset for [data] assignments: param_col + ARROW_OFFSET → '←' column
ARROW_OFFSET = 6

# DrawingML namespaces
NSMAP_XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
NSMAP_A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'
NSMAP_R   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

REQUEST_COLUMN_WIDTHS = {
    'A': 10.1796875, 'B': 10.36328125, 'C': 27.0, 'D': 34.36328125,
    'E': 36.36328125, 'F': 31.453125, 'G': 6.0, 'H': 10.0,
    'I': 13.0, 'J': 49.0, 'K': 3.1796875, 'L': 11.81640625, 'M': 11.81640625,
}

RESPONSE_COLUMN_WIDTHS = {
    'A': 10.1796875, 'B': 11.453125, 'C': 2.81640625, 'D': 13.0, 'E': 13.0,
    'F': 13.0, 'G': 13.0, 'H': 13.0, 'I': 13.0, 'J': 13.0,
    'K': 34.36328125, 'L': 36.36328125, 'M': 31.453125, 'N': 6.0,
    'O': 10.0, 'P': 13.0, 'Q': 35.453125, 'R': 3.1796875,
    'S': 11.81640625, 'T': 11.81640625,
}

SQL_CLAUSE_KEYWORDS = [
    'DELETE FROM',
    'INSERT INTO',
    'ORDER BY',
    'GROUP BY',
    'INNER JOIN',
    'LEFT JOIN',
    'RIGHT JOIN',
    'FULL JOIN',
    'CROSS JOIN',
    'UNION ALL',
    'SELECT',
    'UPDATE',
    'DELETE',
    'VALUES',
    'WHERE',
    'HAVING',
    'RETURNING',
    'UNION',
    'LIMIT',
    'OFFSET',
    'FROM',
    'JOIN',
    'SET',
    'AND',
    'OR',
    'ON',
    'CASE',
    'WHEN',
    'ELSE',
    'END',
]

FALLBACK_SQL_TABLE_LABELS = {
    'syain_applications': '社員申請',
    'applications': '申請',
    'application_groups': '申請グループ',
    'syain_application_item_values': '社員申請項目値',
    'syain_application_item_histories': '社員申請項目変更履歴',
    'syain_application_item_multiple_histories': '社員申請項目マルチ詳細',
    'syain_application_item_multiples': '社員申請項目マルチ',
    'syain_application_fix_item_values': '社員申請固定項目値',
    'syain_application_approvals': '社員申請承認詳細',
    'syain_application_templates': '社員申請テンプレート',
    'syain_items': '社員項目',
    'taikei_item_master': '体系項目',
    'taikei_total_item_master': '体系合計項目',
}

FALLBACK_SQL_COLUMN_LABELS = {
    'syain_applicationid': '社員申請ID',
    'applicationid': '申請ID',
    'exec_type': '処理区分',
    'kaisyacd': '会社コード',
    'value': '値',
    'itemid': '項目ID',
    'fileid': 'ファイルID',
    'syainseq': '社員連番',
    'syain_itemid': '社員項目ID',
    'taikei_code': '体系コード',
    'taikei_item_code': '体系項目コード',
    'taikei_total_item_code': '体系合計項目コード',
    'total_unit': '集計単位',
    'data_type': 'データ型',
    'fromymd': '有効開始日',
    'toymd': '有効終了日',
    'groupid': 'グループID',
}


def copy_cell_style(src, tgt):
    """Copy cell style from src to tgt."""
    tgt.font = copy(src.font)
    tgt.border = copy(src.border)
    tgt.fill = copy(src.fill)
    tgt.number_format = src.number_format
    tgt.alignment = copy(src.alignment)


def copy_row_style(ws, src_row, dest_row, start_col=1, end_col=20):
    """Copy row style and height from one row to another."""
    for col_idx in range(start_col, end_col + 1):
        src = ws.cell(row=src_row, column=col_idx)
        tgt = ws.cell(row=dest_row, column=col_idx)
        copy_cell_style(src, tgt)
    ws.row_dimensions[dest_row].height = ws.row_dimensions[src_row].height


def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', '_', name).strip()


def _contains_japanese_text(value: str) -> bool:
    return bool(JAPANESE_CHAR_RE.search(value or ''))


def _find_disallowed_simplified_phrases(value: str) -> list[str]:
    text = value or ''
    return [phrase for phrase in DISALLOWED_SIMPLIFIED_PHRASES if phrase in text]


def validate_japanese_content(config: dict):
    """Fail fast when human-facing prose is not written in Japanese."""
    violations = []

    def check_text(path: str, value, *, required=False):
        if value is None:
            return
        if not isinstance(value, str):
            return
        text = value.strip()
        if not text:
            return

        disallowed = _find_disallowed_simplified_phrases(text)
        if disallowed:
            violations.append(
                f'{path}: 日本語文言に使用しない簡体字語句を含みます ({",".join(disallowed)})'
            )

        if required and not _contains_japanese_text(text):
            violations.append(f'{path}: 日本語文言が必須です')

    cover = config.get('cover', {}) or {}
    api_info = config.get('api_info', {}) or {}
    overview = config.get('overview', {}) or {}
    sequence = config.get('sequence', {}) or {}
    processing_detail = config.get('processing_detail', {}) or {}

    check_text('cover.api_name', cover.get('api_name'), required=True)
    check_text('api_info.description', api_info.get('description'), required=True)
    check_text('api_info.response_note', api_info.get('response_note'), required=False)
    check_text('overview.summary', overview.get('summary'), required=True)

    for idx, step in enumerate(overview.get('flow_steps', []) or [], 1):
        if isinstance(step, dict):
            check_text(f'overview.flow_steps[{idx}].label', step.get('label'), required=True)

    for idx, param in enumerate(config.get('request_params', []) or [], 1):
        if not isinstance(param, dict):
            continue
        check_text(f'request_params[{idx}].description', param.get('description'), required=True)
        check_text(f'request_params[{idx}].note', param.get('note'), required=False)

    for idx, param in enumerate(config.get('response_params', []) or [], 1):
        if not isinstance(param, dict):
            continue
        check_text(f'response_params[{idx}].description', param.get('description'), required=True)
        check_text(f'response_params[{idx}].note', param.get('note'), required=False)

    check_text('sequence.client_component', sequence.get('client_component'), required=True)
    check_text('sequence.api_title', sequence.get('api_title'), required=True)
    for idx, step in enumerate(sequence.get('steps', []) or [], 1):
        if not isinstance(step, dict):
            continue
        check_text(f'sequence.steps[{idx}].trigger', step.get('trigger'), required=False)
        check_text(f'sequence.steps[{idx}].description', step.get('description'), required=True)
        check_text(f'sequence.steps[{idx}].status', step.get('status'), required=False)
        for detail_idx, detail in enumerate(step.get('details', []) or [], 1):
            check_text(
                f'sequence.steps[{idx}].details[{detail_idx}]',
                detail,
                required=True,
            )

    for idx, step in enumerate(processing_detail.get('steps', []) or [], 1):
        if not isinstance(step, dict):
            continue
        check_text(f'processing_detail.steps[{idx}].title', step.get('title'), required=True)
        for content_idx, content in enumerate(step.get('content', []) or [], 1):
            if isinstance(content, str):
                check_text(
                    f'processing_detail.steps[{idx}].content[{content_idx}]',
                    content,
                    required=True,
                )
                continue
            if not isinstance(content, dict):
                continue
            content_type = content.get('type')
            if content_type in {'sql', 'sql_structured', 'mybatis_sql', 'call', 'assign'}:
                continue
            if content_type == 'sub':
                check_text(
                    f'processing_detail.steps[{idx}].content[{content_idx}].text',
                    content.get('text'),
                    required=True,
                )
            else:
                check_text(
                    f'processing_detail.steps[{idx}].content[{content_idx}].text',
                    content.get('text'),
                    required=True,
                )

    if violations:
        joined = '\n'.join(f'  - {item}' for item in violations)
        raise ValueError(
            '日本語必須チェックに失敗しました。以下の文言を日本語へ修正してください。\n'
            f'{joined}'
        )


def normalize_match_text(value: str) -> str:
    if value is None:
        return ''
    text = str(value).lower()
    text = re.sub(r'[\s\u3000_\-./()（）]+', '', text)
    return text


def build_output_name_from_progress_workbook(config: dict) -> str | None:
    progress = config.get('progress_lookup', {}) or {}
    workbook_path = progress.get('workbook') or config.get('progress_workbook')
    if not workbook_path:
        return None

    workbook_file = Path(workbook_path)
    if not workbook_file.is_absolute():
        workbook_file = Path.cwd() / workbook_file
    if not workbook_file.exists():
        return None

    feature_id_hint = (
        progress.get('feature_id')
        or config.get('feature_id')
        or config.get('screen', {}).get('id')
        or ''
    )
    detail_id_hint = (
        progress.get('detail_id')
        or config.get('detail_id')
        or config.get('api_info', {}).get('operation_id')
        or ''
    )
    detail_title_hint = (
        progress.get('detail_title')
        or config.get('cover', {}).get('api_name')
        or config.get('api_info', {}).get('description')
        or ''
    )

    feature_id_norm = normalize_match_text(feature_id_hint)
    detail_id_norm = normalize_match_text(detail_id_hint)
    detail_title_norm = normalize_match_text(detail_title_hint)

    wb = openpyxl.load_workbook(workbook_file, data_only=True)
    best = None

    try:
        for sheet_name in wb.sheetnames:
            if sheet_name == '改定履歴':
                continue
            ws = wb[sheet_name]
            current_feature_name = ''
            current_feature_id = ''
            feature_matched = not feature_id_norm

            for row_idx in range(1, ws.max_row + 1):
                feature_name = ws.cell(row_idx, 2).value
                feature_id = ws.cell(row_idx, 3).value
                detail_name = ws.cell(row_idx, 4).value
                detail_id = ws.cell(row_idx, 5).value

                if feature_name and feature_id:
                    current_feature_name = str(feature_name).strip()
                    current_feature_id = str(feature_id).strip()
                    feature_matched = (
                        not feature_id_norm
                        or normalize_match_text(current_feature_id) == feature_id_norm
                    )

                if not feature_matched or not detail_name:
                    continue

                score = 0
                detail_name_norm = normalize_match_text(detail_name)
                detail_id_norm_row = normalize_match_text(detail_id)

                if detail_id_norm and detail_id_norm == detail_id_norm_row:
                    score += 100
                if detail_title_norm and detail_title_norm in detail_name_norm:
                    score += 50
                if detail_title_norm:
                    title_tokens = [tok for tok in re.split(r'[・_＿\-]', str(detail_title_hint)) if tok]
                    for token in title_tokens:
                        token_norm = normalize_match_text(token)
                        if token_norm and token_norm in detail_name_norm:
                            score += 10

                if score <= 0:
                    continue

                candidate = {
                    'score': score,
                    'feature_name': current_feature_name,
                    'feature_id': current_feature_id,
                    'detail_name': str(detail_name).strip(),
                    'detail_id': str(detail_id).strip() if detail_id else '',
                    'sheet_name': sheet_name,
                    'row': row_idx,
                }
                if best is None or candidate['score'] > best['score']:
                    best = candidate
    finally:
        wb.close()

    if not best:
        return None

    feature_name = best['feature_name']
    detail_name = re.sub(r'\.xlsx$', '', best['detail_name'], flags=re.IGNORECASE)
    m = re.match(r'^(\d+)\.(.+)$', feature_name)
    if m:
        seq_no = m.group(1)
        if detail_name.startswith(f'{seq_no}.'):
            return sanitize_filename(f'{detail_name}.xlsx')
        return sanitize_filename(f'{seq_no}.{detail_name}.xlsx')

    return sanitize_filename(f'{detail_name}.xlsx')


def derive_operation_label_from_api_id(api_id: str) -> str:
    last = api_id.split('/')[-1] if api_id else '表示'
    lowered = last.lower()
    if 'showblock' in lowered or 'show' in lowered or lowered in {'get', 'detail'}:
        return '表示'
    if 'init' in lowered or 'reload' in lowered:
        return '初期表示'
    if 'search' in lowered or 'list' in lowered:
        return '検索'
    if 'save' in lowered or 'regist' in lowered or 'create' in lowered:
        return '登録'
    if 'update' in lowered:
        return '更新'
    if 'delete' in lowered or 'remove' in lowered:
        return '削除'
    if 'setting' in lowered:
        return '設定'
    return last


def derive_feature_name(config: dict) -> str:
    cover = config.get('cover', {}) or {}
    api_info = config.get('api_info', {}) or {}
    return (
        cover.get('feature_name')
        or cover.get('function_name')
        or api_info.get('feature_name')
        or ''
    ).strip()


def derive_operation_name(config: dict) -> str:
    cover = config.get('cover', {}) or {}
    api_info = config.get('api_info', {}) or {}
    explicit = (
        cover.get('operation_name')
        or api_info.get('operation_name')
        or ''
    ).strip()
    if explicit:
        return explicit

    api_id = (cover.get('api_id') or '').strip()
    if api_id:
        return derive_operation_label_from_api_id(api_id)
    return ''


def build_default_output_name(config: dict) -> str:
    progress_name = build_output_name_from_progress_workbook(config)
    if progress_name:
        return str((Path.cwd() / 'output' / progress_name).resolve())

    cover = config.get('cover', {})
    spec_no = (
        cover.get('spec_no')
        or cover.get('document_no')
        or cover.get('number')
        or ''
    )
    feature_name = derive_feature_name(config)
    operation_name = derive_operation_name(config)
    api_name = cover.get('api_name') or feature_name or 'API'
    name_parts = ['API設計書']
    if feature_name:
        name_parts.append(feature_name)
    if operation_name and operation_name != feature_name:
        name_parts.append(operation_name)
    elif not feature_name:
        name_parts.append(api_name)
    safe_api_name = sanitize_filename('-'.join(part for part in name_parts if part))
    prefix = str(spec_no).strip()
    if prefix:
        prefix = prefix.rstrip('.')
        filename = f'{prefix}.{safe_api_name}.xlsx'
    else:
        filename = f'{safe_api_name}.xlsx'
    return str((Path.cwd() / 'output' / filename).resolve())


# ===================================================================
# DrawingML XML Generators
# ===================================================================
def _make_flowchart_xml(steps):
    """Generate DrawingML XML for a flowchart (処理概要図).

    Creates editable Excel shapes (rectangles + rounded rectangles + connectors).
    Returns XML string for a wsDr element.
    """
    def _normalize_step(step):
        if isinstance(step, dict):
            label = step.get('label') or step.get('text') or ''
            step_type = step.get('type', 'process')
            geom = 'roundRect' if step_type == 'terminal' else 'rect'
            is_terminal = step_type == 'terminal'
            return (label, geom, is_terminal)
        return (str(step), 'rect', False)

    # Build items: start + steps + end
    items = [('開始', 'roundRect', True)] + \
            [_normalize_step(s) for s in steps] + \
            [('終了', 'roundRect', True)]

    n = len(items)
    shapes_xml = []
    shape_id = 2

    # Layout constants (EMU units: 1cm = 360000 EMU)
    box_w = 2700000     # ~7.5cm
    box_h = 350000      # ~1cm
    x_off = 600000      # left margin
    y_start = 2800000   # start Y (below header area)
    y_spacing = 550000  # spacing between boxes

    # Fill colors
    fill_terminal = '<a:solidFill><a:srgbClr val="C8E6C9"/></a:solidFill>'
    fill_process  = '<a:solidFill><a:srgbClr val="BBDEFB"/></a:solidFill>'

    for i, (text, geom, is_terminal) in enumerate(items):
        y = y_start + i * y_spacing
        fill = fill_terminal if is_terminal else fill_process

        shape_xml = f'''<xdr:oneCellAnchor>
  <xdr:from><xdr:col>1</xdr:col><xdr:colOff>100000</xdr:colOff><xdr:row>{12 + i * 3}</xdr:row><xdr:rowOff>20000</xdr:rowOff></xdr:from>
  <xdr:ext cx="{box_w}" cy="{box_h}"/>
  <xdr:sp macro="" textlink="">
    <xdr:nvSpPr>
      <xdr:cNvPr id="{shape_id}" name="Shape {shape_id}"/>
      <xdr:cNvSpPr/>
    </xdr:nvSpPr>
    <xdr:spPr>
      <a:xfrm><a:off x="{x_off}" y="{y}"/><a:ext cx="{box_w}" cy="{box_h}"/></a:xfrm>
      <a:prstGeom prst="{geom}"><a:avLst/></a:prstGeom>
      {fill}
      <a:ln w="9525"><a:solidFill><a:srgbClr val="555555"/></a:solidFill></a:ln>
    </xdr:spPr>
    <xdr:txBody>
      <a:bodyPr vertOverflow="clip" wrap="square" lIns="36000" tIns="36000" rIns="36000" bIns="36000" anchor="ctr"/>
      <a:lstStyle/>
      <a:p>
        <a:pPr algn="ctr"/>
        <a:r>
          <a:rPr lang="ja-JP" sz="900">
            <a:latin typeface="Kosugi"/>
            <a:ea typeface="Kosugi"/>
          </a:rPr>
          <a:t>{text}</a:t>
        </a:r>
      </a:p>
    </xdr:txBody>
  </xdr:sp>
  <xdr:clientData/>
</xdr:oneCellAnchor>'''
        shapes_xml.append(shape_xml)
        shape_id += 1

    # Add arrow connectors between boxes
    for i in range(n - 1):
        y1 = y_start + i * y_spacing + box_h
        y2 = y_start + (i + 1) * y_spacing
        cx = x_off + box_w // 2
        arrow_h = y2 - y1

        row_from = 12 + i * 3 + 2  # approximate row for bottom of box
        row_to = 12 + (i + 1) * 3  # approximate row for top of next box

        connector_xml = f'''<xdr:twoCellAnchor>
  <xdr:from><xdr:col>3</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>{row_from}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
  <xdr:to><xdr:col>3</xdr:col><xdr:colOff>10000</xdr:colOff><xdr:row>{row_to}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>
  <xdr:cxnSp macro="">
    <xdr:nvCxnSpPr>
      <xdr:cNvPr id="{shape_id}" name="Arrow {shape_id}"/>
      <xdr:cNvCxnSpPr/>
    </xdr:nvCxnSpPr>
    <xdr:spPr>
      <a:xfrm><a:off x="{cx}" y="{y1}"/><a:ext cx="0" cy="{arrow_h}"/></a:xfrm>
      <a:prstGeom prst="straightConnector1"><a:avLst/></a:prstGeom>
      <a:ln w="9525">
        <a:solidFill><a:srgbClr val="333333"/></a:solidFill>
        <a:tailEnd type="arrow"/>
      </a:ln>
    </xdr:spPr>
  </xdr:cxnSp>
  <xdr:clientData/>
</xdr:twoCellAnchor>'''
        shapes_xml.append(connector_xml)
        shape_id += 1

    xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="{NSMAP_XDR}" xmlns:a="{NSMAP_A}">
{''.join(shapes_xml)}
</xdr:wsDr>'''
    return xml


def _make_sequence_arrows_xml(request_rows, response_rows):
    """Generate DrawingML XML for APIシーケンス&DFD arrow connectors.

    request_rows: list of row numbers where request arrows should be
    response_rows: list of row numbers where response arrows should be
    """
    shapes_xml = []
    shape_id = 2

    # Request arrow: → from col C to col D (client → backend)
    for row in request_rows:
        shapes_xml.append(f'''<xdr:twoCellAnchor>
  <xdr:from><xdr:col>3</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>{row - 1}</xdr:row><xdr:rowOff>66675</xdr:rowOff></xdr:from>
  <xdr:to><xdr:col>3</xdr:col><xdr:colOff>277628</xdr:colOff><xdr:row>{row - 1}</xdr:row><xdr:rowOff>67652</xdr:rowOff></xdr:to>
  <xdr:cxnSp macro="">
    <xdr:nvCxnSpPr>
      <xdr:cNvPr id="{shape_id}" name="Arrow {shape_id}"/>
      <xdr:cNvCxnSpPr/>
    </xdr:nvCxnSpPr>
    <xdr:spPr>
      <a:xfrm flipV="1"><a:off x="3705225" y="{1647825 + (row - 10) * 162000}"/><a:ext cx="277628" cy="977"/></a:xfrm>
      <a:prstGeom prst="straightConnector1"><a:avLst/></a:prstGeom>
      <a:ln w="12700"><a:solidFill><a:schemeClr val="tx1"/></a:solidFill><a:tailEnd type="arrow"/></a:ln>
    </xdr:spPr>
    <xdr:style>
      <a:lnRef idx="1"><a:schemeClr val="accent1"/></a:lnRef>
      <a:fillRef idx="0"><a:schemeClr val="accent1"/></a:fillRef>
      <a:effectRef idx="0"><a:schemeClr val="accent1"/></a:effectRef>
      <a:fontRef idx="minor"><a:schemeClr val="tx1"/></a:fontRef>
    </xdr:style>
  </xdr:cxnSp>
  <xdr:clientData/>
</xdr:twoCellAnchor>''')
        shape_id += 1

    # Response arrow: ← from col D to col C (backend → client)
    for row in response_rows:
        shapes_xml.append(f'''<xdr:twoCellAnchor>
  <xdr:from><xdr:col>2</xdr:col><xdr:colOff>1638300</xdr:colOff><xdr:row>{row - 1}</xdr:row><xdr:rowOff>76200</xdr:rowOff></xdr:from>
  <xdr:to><xdr:col>3</xdr:col><xdr:colOff>267626</xdr:colOff><xdr:row>{row - 1}</xdr:row><xdr:rowOff>77177</xdr:rowOff></xdr:to>
  <xdr:cxnSp macro="">
    <xdr:nvCxnSpPr>
      <xdr:cNvPr id="{shape_id}" name="Arrow {shape_id}"/>
      <xdr:cNvCxnSpPr/>
    </xdr:nvCxnSpPr>
    <xdr:spPr>
      <a:xfrm flipV="1"><a:off x="3695700" y="{1647825 + (row - 10) * 162000}"/><a:ext cx="277151" cy="977"/></a:xfrm>
      <a:prstGeom prst="straightConnector1"><a:avLst/></a:prstGeom>
      <a:ln w="12700"><a:solidFill><a:schemeClr val="tx1"/></a:solidFill><a:headEnd type="arrow"/><a:tailEnd type="none"/></a:ln>
    </xdr:spPr>
    <xdr:style>
      <a:lnRef idx="1"><a:schemeClr val="accent1"/></a:lnRef>
      <a:fillRef idx="0"><a:schemeClr val="accent1"/></a:fillRef>
      <a:effectRef idx="0"><a:schemeClr val="accent1"/></a:effectRef>
      <a:fontRef idx="minor"><a:schemeClr val="tx1"/></a:fontRef>
    </xdr:style>
  </xdr:cxnSp>
  <xdr:clientData/>
</xdr:twoCellAnchor>''')
        shape_id += 1

    xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="{NSMAP_XDR}" xmlns:a="{NSMAP_A}">
{''.join(shapes_xml)}
</xdr:wsDr>'''
    return xml


def inject_drawings_into_xlsx(xlsx_path, sheet_drawings):
    """Post-process xlsx to inject DrawingML drawings.

    sheet_drawings: dict of {sheet_name: drawing_xml_string}

    Rewrites the xlsx zip to add drawing XML files and wire them
    to the correct worksheet via relationships.
    """
    import html

    tmp_path = xlsx_path + '.tmp'

    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        existing_files = zin.namelist()

        # Read workbook.xml to map sheet names → sheet file paths
        workbook_xml = zin.read('xl/workbook.xml').decode('utf-8')
        workbook_rels = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')

        # Parse: rId → target file (handle both possible attribute orders)
        rid_to_target = {}
        # Pattern 1: Target before Id
        for m in re.finditer(r'<Relationship[^>]*Target="([^"]+)"[^>]*Id="([^"]+)"', workbook_rels):
            target, rid = m.groups()
            target = target.lstrip('/')
            if not target.startswith('xl/'):
                target = 'xl/' + target
            rid_to_target[rid] = target
        # Pattern 2: Id before Target
        for m in re.finditer(r'<Relationship[^>]*Id="([^"]+)"[^>]*Target="([^"]+)"', workbook_rels):
            rid, target = m.groups()
            target = target.lstrip('/')
            if not target.startswith('xl/'):
                target = 'xl/' + target
            rid_to_target[rid] = target

        # Parse: sheet name → rId (handle XML-escaped names like &amp;)
        sheet_files = {}
        for m in re.finditer(r'<sheet\s[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', workbook_xml):
            sname_escaped, rid = m.groups()
            sname = html.unescape(sname_escaped)
            if rid in rid_to_target:
                sheet_files[sname] = rid_to_target[rid]

        print(f'  [DrawingML] Sheet→File mapping: {sheet_files}')

        # Determine next drawing number
        existing_drawings = [f for f in existing_files if re.match(r'xl/drawings/drawing\d+\.xml', f)]
        next_num = len(existing_drawings) + 1

        # Collect all modifications
        new_files = {}          # path → content to add
        modified_files = {}     # path → modified content
        rels_to_create = {}     # rels path → content
        drawing_nums = []       # tracking drawing numbers added

        for sheet_name, drawing_xml in sheet_drawings.items():
            if sheet_name not in sheet_files:
                print(f'  [DrawingML] WARNING: Sheet "{sheet_name}" not found, skipping')
                continue

            sheet_file = sheet_files[sheet_name]
            drawing_file = f'xl/drawings/drawing{next_num}.xml'
            drawing_rel_id = f'rIdDraw{next_num}'

            # 1. Add drawing XML file
            new_files[drawing_file] = drawing_xml

            # 2. Modify sheet XML to add <drawing> element
            sheet_xml = zin.read(sheet_file).decode('utf-8')
            if '<drawing ' not in sheet_xml:
                sheet_xml = sheet_xml.replace(
                    '</worksheet>',
                    f'<drawing r:id="{drawing_rel_id}"/></worksheet>'
                )
                if 'xmlns:r=' not in sheet_xml:
                    sheet_xml = sheet_xml.replace(
                        '<worksheet',
                        f'<worksheet xmlns:r="{NSMAP_R}"',
                        1
                    )
            modified_files[sheet_file] = sheet_xml

            # 3. Create/update sheet .rels file
            sheet_dir = os.path.dirname(sheet_file)
            sheet_base = os.path.basename(sheet_file)
            rels_path = f'{sheet_dir}/_rels/{sheet_base}.rels'

            if rels_path in existing_files:
                rels_xml = zin.read(rels_path).decode('utf-8')
                rels_xml = rels_xml.replace(
                    '</Relationships>',
                    f'<Relationship Id="{drawing_rel_id}" '
                    f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
                    f'Target="../drawings/drawing{next_num}.xml"/>'
                    f'</Relationships>'
                )
            else:
                rels_xml = (
                    f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    f'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    f'<Relationship Id="{drawing_rel_id}" '
                    f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
                    f'Target="../drawings/drawing{next_num}.xml"/>'
                    f'</Relationships>'
                )
            rels_to_create[rels_path] = rels_xml

            drawing_nums.append(next_num)
            next_num += 1

        # Write new zip
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            written = set()

            for item in existing_files:
                if item in modified_files:
                    zout.writestr(item, modified_files[item])
                elif item in rels_to_create:
                    zout.writestr(item, rels_to_create[item])
                elif item == '[Content_Types].xml':
                    ct = zin.read(item).decode('utf-8')
                    for dnum in drawing_nums:
                        ct_entry = (f'<Override PartName="/xl/drawings/drawing{dnum}.xml" '
                                    f'ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>')
                        if ct_entry not in ct:
                            ct = ct.replace('</Types>', f'{ct_entry}</Types>')
                    zout.writestr(item, ct)
                else:
                    zout.writestr(item, zin.read(item))
                written.add(item)

            # Write new rels files (if not already existing in zip)
            for rp, rc in rels_to_create.items():
                if rp not in written:
                    zout.writestr(rp, rc)

            # Write new drawing files
            for dp, dc in new_files.items():
                zout.writestr(dp, dc)

    # Replace original with tmp
    shutil.move(tmp_path, xlsx_path)
    print(f'  [ok] DrawingML注入完了 ({len(sheet_drawings)}シート, {len(drawing_nums)}図面)')


# ===================================================================
# Generator Class
# ===================================================================
class ApiSpecGenerator:
    """API設計書を1エンドポイント単位で生成するジェネレーター。"""

    def __init__(self, config: dict, template_path: str):
        self.config = config
        self.wb = openpyxl.load_workbook(template_path)
        self._style_reference_wb = self._load_style_reference_workbook(template_path)
        self._sql_translation = self._build_sql_translation()
        self._ddl_table_labels = self._load_ddl_table_labels()
        # Track arrow positions for post-processing
        self._seq_request_rows = []
        self._seq_response_rows = []

    def _build_sql_translation(self):
        custom = self.config.get('sql_translation', {})
        return {
            'tables': {**FALLBACK_SQL_TABLE_LABELS, **custom.get('tables', {})},
            'columns': {**FALLBACK_SQL_COLUMN_LABELS, **custom.get('columns', {})},
            'params': dict(custom.get('params', {})),
            'constants': dict(custom.get('constants', {})),
            'comments': dict(custom.get('comments', {})),
        }

    def _load_style_reference_workbook(self, template_path: str):
        """Load a nearby high-quality API workbook as a sheet layout donor if available."""
        explicit = self.config.get('style_reference_workbook')
        candidates = []
        if explicit:
            candidates.append(Path(explicit))

        tpl_path = Path(template_path).resolve()
        workspace_root = tpl_path.parent.parent.parent
        ref_dir = workspace_root / 'reference'
        if ref_dir.exists():
            candidates.extend(sorted(ref_dir.glob('22.API設計書-*.xlsx')))

        for candidate in candidates:
            try:
                if candidate.exists():
                    print(f'  [ok] レイアウト参照ブック: {candidate}')
                    return openpyxl.load_workbook(candidate)
            except Exception as exc:
                print(f'  [warn] 参照ブックを読めません: {candidate} ({exc})')
        return None

    def _iter_candidate_ddl_roots(self):
        seen = set()

        explicit_roots = self.config.get('ddl_roots', []) or []
        if isinstance(explicit_roots, str):
            explicit_roots = [explicit_roots]

        for root in explicit_roots:
            path = Path(root)
            if path.exists():
                resolved = path.resolve()
                if resolved not in seen:
                    seen.add(resolved)
                    yield resolved

        mapper_paths = set()
        mapper_xml = self.config.get('mapper_xml')
        if mapper_xml:
            mapper_paths.add(mapper_xml)

        for item in self._walk_processing_detail_items():
            if isinstance(item, dict) and item.get('type') == 'mybatis_sql':
                xml_path = item.get('mapper_xml')
                if xml_path:
                    mapper_paths.add(xml_path)

        for xml_path in mapper_paths:
            path = Path(xml_path)
            if not path.exists():
                continue
            for parent in path.parents:
                if (parent / 'db' / 'createtable').exists() or (parent / 'db' / 'createview').exists():
                    resolved = (parent / 'db').resolve()
                    if resolved not in seen:
                        seen.add(resolved)
                        yield resolved
                    break

    def _load_ddl_table_labels(self):
        labels = {}
        for db_root in self._iter_candidate_ddl_roots():
            for folder_name in ('createtable', 'createview'):
                folder = db_root / folder_name
                if not folder.exists():
                    continue
                for sql_file in folder.glob('Crt_*.sql'):
                    try:
                        text = sql_file.read_text(encoding='utf-8')
                    except UnicodeDecodeError:
                        text = sql_file.read_text(encoding='utf-8', errors='ignore')

                    physical_match = re.search(
                        r'create\s+(?:table|view)\s+`?([A-Za-z_][A-Za-z0-9_]*)`?',
                        text,
                        re.IGNORECASE,
                    )
                    if not physical_match:
                        continue
                    physical_name = physical_match.group(1)

                    logical_name = ''
                    table_name_match = re.search(r'TableName:\s*(.+)', text, re.IGNORECASE)
                    if table_name_match:
                        logical_name = table_name_match.group(1).strip()
                    if not logical_name:
                        comment_match = re.search(r'COMMENT\s*=\s*\'([^\']+)\'', text, re.IGNORECASE)
                        if comment_match:
                            logical_name = comment_match.group(1).strip()
                    if not logical_name:
                        continue

                    labels.setdefault(physical_name, logical_name)
        return labels

    def _format_sequence_table_name(self, table_name):
        text = str(table_name or '').strip()
        if not text or re.search(r'（.*）|\(.*\)', text):
            return text

        logical_name = (
            self._ddl_table_labels.get(text)
            or self._sql_translation.get('tables', {}).get(text)
        )
        if logical_name and logical_name != text:
            return f'{logical_name}({text})'
        return text

    def _write_detail_text_cell(self, ws, cell_ref, value, *, horizontal='left'):
        ws[cell_ref] = value
        ws[cell_ref].font = FONT_NORMAL
        ws[cell_ref].alignment = Alignment(
            vertical='center',
            horizontal=horizontal,
            wrap_text=False,
            shrink_to_fit=False,
        )

    # ------------------------------------------------------------------
    # 表紙
    # ------------------------------------------------------------------
    def fill_cover(self):
        ws = self.wb['表紙']
        cover = self.config.get('cover', {})
        company = cover.get('company', '')
        mapping = {
            'C4': company,
            'C5': cover.get('project', ''),
            'C6': cover.get('system', ''),
            'C7': cover.get('api_name', ''),
            'C8': cover.get('api_id', ''),
            # C9: 原版は空白。api_url が明示指定された場合のみ書き込む
            'C11': cover.get('create_date', ''),
            'C12': cover.get('author', ''),
            'C13': cover.get('update_date', datetime.now().strftime('%Y-%m-%d')),
            'C14': cover.get('update_author', cover.get('author', '')),
            # D16: コピーライト表示 = "©" + 会社名（テンプレートの固定値を上書き）
            'D16': f'©{company}' if company else '',
        }
        # C9 は api_url が明示的に指定された場合のみ
        api_url = cover.get('api_url')
        if api_url:
            mapping['C9'] = api_url
        for addr, val in mapping.items():
            ws[addr] = val

    # ------------------------------------------------------------------
    # リクエストAPIパラメーター / レスポンスAPIパラメーター
    # ------------------------------------------------------------------
    def fill_request_params(self):
        ws = self.wb['リクエストAPIパラメーター']
        api = self.config.get('api_info', {})
        cover = self.config.get('cover', {})

        donor_max_row = self._prepare_request_param_sheet(ws)
        ws['A8'] = api.get('method', 'POST')
        ws['B8'] = api.get('url', '').lstrip('/')
        ws['J8'] = api.get('description', '')
        ws['L8'] = cover.get('update_date', cover.get('create_date', ''))
        ws['M8'] = cover.get('update_author', cover.get('author', ''))
        params = self.config.get('request_params', [])
        flat = self._flatten_params(params)
        end_row = 12 + max(len(flat) - 1, 0)
        self._ensure_sheet_row_capacity(
            ws,
            start_row=12,
            end_row=end_row,
            template_row=12,
            max_col=20,
            donor_max_row=donor_max_row,
        )
        self._write_request_param_table(ws, params, start_row=12)

    def fill_response_params(self):
        ws = self.wb['レスポンスAPIパラメーター']
        cover = self.config.get('cover', {})
        api = self.config.get('api_info', {})

        donor_max_row = self._prepare_response_param_sheet(ws)
        ws['A8'] = api.get('content_type', 'application/json')
        ws['Q8'] = api.get('response_note', '')
        ws['S8'] = cover.get('update_date', cover.get('create_date', ''))
        ws['T8'] = cover.get('update_author', cover.get('author', ''))
        response_label = api.get('response_label', self.config.get('response_label', ''))
        if response_label:
            ws['B12'] = response_label
        params = self.config.get('response_params', [])
        flat = self._flatten_params(params)
        end_row = 13 + max(len(flat) - 1, 0)
        self._ensure_sheet_row_capacity(
            ws,
            start_row=13,
            end_row=end_row,
            template_row=13,
            max_col=20,
            donor_max_row=donor_max_row,
        )
        self._write_response_param_table(ws, params, start_row=13)
        self._normalize_response_param_rows(ws, start_row=13, end_row=end_row)

    def _prepare_request_param_sheet(self, ws):
        """Align request sheet to the reference workbook layout."""
        donor_max_row = ws.max_row
        if self._style_reference_wb and 'リクエストAPIパラメーター' in self._style_reference_wb.sheetnames:
            donor = self._style_reference_wb['リクエストAPIパラメーター']
            self._clone_sheet_layout(donor, ws)
            donor_max_row = donor.max_row
            self._clear_sheet_values(ws, start_row=8, end_row=8, max_col=20)
            self._clear_sheet_values(ws, start_row=12, end_row=max(ws.max_row, donor.max_row), max_col=20)
            self._unmerge_ranges_from_row(ws, start_row=12)
        else:
            self._reset_sheet_merges(ws, [
                'B7:I7',
                'B8:I8',
                'B9:I9',
            ])
            self._apply_column_widths(ws, REQUEST_COLUMN_WIDTHS)
            copy_row_style(ws, 10, 11)
            header_base = ws['C10']
            for cell_ref in ['D11', 'E11', 'F11', 'G11', 'H11', 'I11', 'J11', 'L11', 'M11']:
                copy_cell_style(header_base, ws[cell_ref])
            header_map = {
                'A7': 'メソッド',
                'B7': 'url',
                'J7': '備考',
                'L7': '改定日付',
                'M7': '改定者',
                'A11': 'チケットNo.',
                'B11': 'No',
                'C11': 'パラメータ名',
                'D11': '項目名',
                'E11': '項目説明',
                'F11': '記述例',
                'G11': '必須',
                'H11': 'データ型',
                'I11': 'データ長',
                'J11': '備考',
                'L11': '改定日付',
                'M11': '改定者',
            }
            for cell_ref, value in header_map.items():
                ws[cell_ref] = value
                ws[cell_ref].font = FONT_HEADER
                ws[cell_ref].alignment = ALIGN_VCC
        return donor_max_row

    def _prepare_response_param_sheet(self, ws):
        """Align response sheet to the reference workbook layout."""
        donor_max_row = ws.max_row
        if self._style_reference_wb and 'レスポンスAPIパラメーター' in self._style_reference_wb.sheetnames:
            donor = self._style_reference_wb['レスポンスAPIパラメーター']
            self._clone_sheet_layout(donor, ws)
            donor_max_row = donor.max_row
            self._clear_sheet_values(ws, start_row=8, end_row=8, max_col=20)
            self._clear_sheet_values(ws, start_row=12, end_row=max(ws.max_row, donor.max_row), max_col=20)
            self._unmerge_ranges_from_row(ws, start_row=13)
        else:
            self._reset_sheet_merges(ws, [
                'A7:K7',
                'A8:K8',
                'A9:K9',
                'C11:J11',
            ])
            self._apply_column_widths(ws, RESPONSE_COLUMN_WIDTHS)
            copy_row_style(ws, 10, 11)
            copy_row_style(ws, 11, 12)
            header_map = {
                'A7': 'コンテントタイプ',
                'Q7': '備考',
                'S7': '改定日付',
                'T7': '改定者',
                'A11': 'チケットNo.',
                'B11': 'レスポンスコード',
                'K11': '項目名',
                'L11': '項目説明',
                'M11': '記述例',
                'N11': '必須',
                'O11': 'データ型',
                'P11': 'データ長',
                'Q11': '備考',
                'S11': '改定日付',
                'T11': '改定者',
            }
            for cell_ref, value in header_map.items():
                ws[cell_ref] = value
                ws[cell_ref].font = FONT_HEADER
                ws[cell_ref].alignment = ALIGN_VCC
            ws['C11'] = 'パラメータ名/階層'
            ws['C11'].font = FONT_HEADER
            ws['C11'].alignment = ALIGN_VCC
        return donor_max_row

    def _clone_sheet_layout(self, donor_ws, target_ws):
        """Clone widths, row heights, merges, formulas, and styles from donor to target."""
        current_merges = [str(rng) for rng in target_ws.merged_cells.ranges]
        for merge in current_merges:
            target_ws.unmerge_cells(merge)

        for col_idx in range(1, max(donor_ws.max_column, target_ws.max_column) + 1):
            col_letter = get_column_letter(col_idx)
            target_ws.column_dimensions[col_letter].width = donor_ws.column_dimensions[col_letter].width

        for row_idx in range(1, donor_ws.max_row + 1):
            target_ws.row_dimensions[row_idx].height = donor_ws.row_dimensions[row_idx].height
            for col_idx in range(1, donor_ws.max_column + 1):
                donor_cell = donor_ws.cell(row=row_idx, column=col_idx)
                target_cell = target_ws.cell(row=row_idx, column=col_idx)
                copy_cell_style(donor_cell, target_cell)
                target_cell.value = donor_cell.value

        for merge in donor_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merge))

    def _clear_sheet_values(self, ws, start_row, end_row, max_col):
        """Clear cell values while keeping styles in the data-entry area."""
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                cell.value = None

    def _unmerge_ranges_from_row(self, ws, start_row):
        """Unmerge data-area ranges so row writers do not hit read-only merged cells."""
        merges = [str(rng) for rng in ws.merged_cells.ranges if rng.min_row >= start_row]
        for merge in merges:
            ws.unmerge_cells(merge)

    def _ensure_sheet_row_capacity(self, ws, *, start_row, end_row, template_row, max_col, donor_max_row):
        """Extend template row styling when generated rows exceed the donor sheet height."""
        if end_row <= donor_max_row:
            return
        for row_idx in range(max(donor_max_row + 1, start_row), end_row + 1):
            copy_row_style(ws, template_row, row_idx, start_col=1, end_col=max_col)

    def _reset_sheet_merges(self, ws, target_merges):
        current_merges = [str(rng) for rng in ws.merged_cells.ranges]
        for merge in current_merges:
            ws.unmerge_cells(merge)
        for merge in target_merges:
            ws.merge_cells(merge)

    def _apply_column_widths(self, ws, width_map):
        for col_letter, width in width_map.items():
            ws.column_dimensions[col_letter].width = width

    def _write_request_param_table(self, ws, params, start_row):
        """Write request parameter rows with the request-sheet column mapping."""
        row = start_row
        flat = self._flatten_params(params)

        for idx, p in enumerate(flat):
            depth = p.get('depth', 0)
            prefix = '　' * depth
            param_name = self._pick_param_value(
                p, 'name', 'param_name', 'field_name', 'api_name', 'item_name'
            )
            display_name = self._pick_param_value(
                p, 'display_name', 'item_name', 'jp_name', 'label', 'item_label'
            )
            if not display_name:
                description = p.get('description', '')
                display_name = description if description else param_name
            if not param_name:
                param_name = display_name

            # No
            ws[f'B{row}'] = idx + 1
            ws[f'B{row}'].font = FONT_NORMAL
            ws[f'B{row}'].alignment = ALIGN_VCC

            ws[f'C{row}'] = prefix + (param_name or '')
            ws[f'C{row}'].font = FONT_NORMAL
            ws[f'C{row}'].alignment = ALIGN_VCL

            ws[f'D{row}'] = display_name or ''
            ws[f'D{row}'].font = FONT_NORMAL
            ws[f'D{row}'].alignment = ALIGN_VCL

            desc = p.get('description', '')
            if desc:
                ws[f'E{row}'] = desc
                ws[f'E{row}'].font = FONT_NORMAL
                ws[f'E{row}'].alignment = ALIGN_VCL

            example = p.get('example', '')
            if example:
                ws[f'F{row}'] = example
                ws[f'F{row}'].font = FONT_NORMAL
                ws[f'F{row}'].alignment = ALIGN_VCL

            ws[f'G{row}'] = p.get('required', '')
            ws[f'G{row}'].font = FONT_NORMAL
            ws[f'G{row}'].alignment = ALIGN_VCC

            ws[f'H{row}'] = p.get('data_type', '')
            ws[f'H{row}'].font = FONT_NORMAL
            ws[f'H{row}'].alignment = ALIGN_VCL

            ws[f'I{row}'] = p.get('data_length', '-')
            ws[f'I{row}'].font = FONT_NORMAL
            ws[f'I{row}'].alignment = ALIGN_VCC

            note = p.get('note', '')
            if note:
                ws[f'J{row}'] = note
                ws[f'J{row}'].font = FONT_NORMAL
                ws[f'J{row}'].alignment = ALIGN_VCL

            # Response code (for response sheet only)
            rc = p.get('response_code', '')
            if rc:
                ws[f'B{row}'] = rc

            row += 1

        print(f'  [ok] パラメーター: {len(flat)}行書き込み')

    def _write_response_param_table(self, ws, params, start_row):
        """Write response parameter rows with hierarchical indentation."""
        row = start_row
        flat = self._flatten_params(params)

        for idx, p in enumerate(flat):
            depth = p.get('depth', 0)
            name_col = get_column_letter(3 + depth)  # C=0, D=1, E=2, ...

            response_code = p.get('response_code')
            if response_code in (None, '') and idx == 0:
                response_code = self.config.get('api_info', {}).get('success_status', 200)
            if response_code not in (None, ''):
                ws[f'B{row}'] = response_code
                ws[f'B{row}'].font = FONT_NORMAL
                ws[f'B{row}'].alignment = ALIGN_VCC

            ws[f'{name_col}{row}'] = self._pick_param_value(
                p, 'name', 'param_name', 'field_name', 'item_name'
            )
            ws[f'{name_col}{row}'].font = FONT_NORMAL
            ws[f'{name_col}{row}'].alignment = ALIGN_VCL

            item_name = self._pick_param_value(
                p, 'item_name', 'display_name', 'jp_name', 'label'
            )
            prefix = '　' * depth
            ws[f'K{row}'] = prefix + (item_name or '')
            ws[f'K{row}'].font = FONT_NORMAL
            ws[f'K{row}'].alignment = ALIGN_VCL

            desc = p.get('description', '')
            if desc:
                ws[f'L{row}'] = desc
                ws[f'L{row}'].font = FONT_NORMAL
                ws[f'L{row}'].alignment = ALIGN_VCL

            example = p.get('example', '')
            if example:
                ws[f'M{row}'] = example
                ws[f'M{row}'].font = FONT_NORMAL
                ws[f'M{row}'].alignment = ALIGN_VCL

            ws[f'N{row}'] = p.get('required', '')
            ws[f'N{row}'].font = FONT_NORMAL
            ws[f'N{row}'].alignment = ALIGN_VCC

            ws[f'O{row}'] = p.get('data_type', '')
            ws[f'O{row}'].font = FONT_NORMAL
            ws[f'O{row}'].alignment = ALIGN_VCL

            ws[f'P{row}'] = p.get('data_length', '-')
            ws[f'P{row}'].font = FONT_NORMAL
            ws[f'P{row}'].alignment = ALIGN_VCC

            note = p.get('note', '')
            if note:
                ws[f'Q{row}'] = note
                ws[f'Q{row}'].font = FONT_NORMAL
                ws[f'Q{row}'].alignment = ALIGN_VCL

            row += 1

        print(f'  [ok] パラメーター: {len(flat)}行書き込み')

    def _find_response_name_col(self, ws, row):
        for col in range(3, 11):
            cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ''):
                return get_column_letter(col)
        return None

    def _has_table_border(self, ws, row):
        for col in ('K', 'L', 'N', 'O', 'P', 'Q'):
            if ws[f'{col}{row}'].border.left.style == 'thin':
                return True
        return False

    def _normalize_response_param_rows(self, ws, start_row, end_row):
        """Reapply stable row templates so dirty donor rows do not leak into output."""
        template_rows = {}
        for row in range(start_row, end_row + 1):
            name_col = self._find_response_name_col(ws, row)
            if not name_col or not self._has_table_border(ws, row):
                continue
            template_rows.setdefault(name_col, row)

        fallback_templates = {
            'C': start_row,
            'D': start_row + 1,
            'E': start_row + 4,
            'F': start_row + 5,
        }
        for name_col, fallback_row in fallback_templates.items():
            if name_col in template_rows:
                continue
            if fallback_row <= end_row:
                template_rows[name_col] = fallback_row

        for row in range(start_row, end_row + 1):
            name_col = self._find_response_name_col(ws, row)
            if not name_col:
                continue
            template_row = template_rows.get(name_col)
            if template_row is None:
                continue
            for col in range(2, 21):
                copy_cell_style(
                    ws.cell(row=template_row, column=col),
                    ws.cell(row=row, column=col),
                )
            ws.row_dimensions[row].height = ws.row_dimensions[template_row].height

    def _pick_param_value(self, param, *keys):
        for key in keys:
            value = param.get(key)
            if value not in (None, ''):
                return value
        return ''

    def _flatten_params(self, params, depth=0):
        """Flatten hierarchical param tree into list with depth info."""
        result = []
        for p in params:
            flat = dict(p)
            explicit_depth = flat.pop('depth', None)
            if explicit_depth is None:
                flat['depth'] = depth
            elif explicit_depth == 0 and depth > 0:
                flat['depth'] = depth
            else:
                flat['depth'] = explicit_depth
            children = flat.pop('children', [])
            result.append(flat)
            if children:
                result.extend(self._flatten_params(children, flat['depth'] + 1))
        return result

    def _get_sql_translation_map(self, item):
        tables = {
            **self._sql_translation.get('tables', {}),
            **item.get('table_labels', {}),
        }
        columns = {
            **self._sql_translation.get('columns', {}),
            **item.get('column_labels', {}),
        }
        params = {
            **self._sql_translation.get('params', {}),
            **item.get('param_labels', {}),
        }
        constants = {
            **self._sql_translation.get('constants', {}),
            **item.get('constant_labels', {}),
        }
        comments = {
            **self._sql_translation.get('comments', {}),
            **item.get('statement_comments', {}),
        }
        return {
            'tables': tables,
            'columns': columns,
            'params': params,
            'constants': constants,
            'comments': comments,
        }

    def _load_mybatis_statement(self, xml_path: str, statement_id: str):
        xml_file = Path(xml_path)
        if not xml_file.exists():
            raise FileNotFoundError(f'MyBatis XML not found: {xml_file}')

        parser = ET.XMLParser(target=ET.TreeBuilder(insert_comments=True))
        root = ET.parse(xml_file, parser=parser).getroot()
        children = list(root)
        for idx, node in enumerate(children):
            if node.attrib.get('id') != statement_id:
                continue

            comment = ''
            prev_idx = idx - 1
            while prev_idx >= 0:
                prev = children[prev_idx]
                if not isinstance(prev.tag, str):
                    comment = (prev.text or '').strip()
                    if comment:
                        break
                elif (prev.text or '').strip():
                    break
                prev_idx -= 1

            sql_text = self._flatten_mybatis_sql_text(node)
            statement_type = str(node.tag).split('}')[-1].lower()
            return {
                'sql_text': sql_text,
                'statement_type': statement_type,
                'comment': comment,
            }

        raise KeyError(f'MyBatis statement id not found: {statement_id}')

    def _flatten_mybatis_sql_text(self, node):
        parts = []
        if node.text:
            parts.append(node.text)

        for child in list(node):
            tag = str(child.tag).split('}')[-1]
            if tag in {'if', 'when', 'otherwise', 'choose', 'trim', 'where', 'set'}:
                parts.append(self._flatten_mybatis_sql_text(child))
            elif tag == 'foreach':
                open_token = child.attrib.get('open', '')
                close_token = child.attrib.get('close', '')
                separator = child.attrib.get('separator', ', ')
                body = self._flatten_mybatis_sql_text(child).strip()
                if body:
                    parts.append(f' {open_token}{body}{close_token} ')
                elif open_token or close_token:
                    parts.append(f' {open_token}{close_token} ')
                if separator and body and separator not in body:
                    parts.append(' ')
            else:
                parts.append(self._flatten_mybatis_sql_text(child))

            if child.tail:
                parts.append(child.tail)

        return ''.join(parts)

    def _split_sql_lines(self, sql_text: str):
        raw_lines = []
        for line in sql_text.splitlines():
            stripped = line.strip()
            if not stripped:
                continue
            stripped = re.sub(r'\s+', ' ', stripped)
            raw_lines.append(stripped)
        return raw_lines

    def _extract_sql_alias_map(self, sql_lines, translation):
        alias_map = {}
        tables = translation['tables']
        pending_table_keyword = ''
        for line in sql_lines:
            normalized = re.sub(r'\s+', ' ', line.strip())
            match = re.match(
                r'^(FROM|INNER JOIN|LEFT JOIN|RIGHT JOIN|FULL JOIN|CROSS JOIN|JOIN)\s+([A-Za-z_][A-Za-z0-9_]*)(?:\s+(?:AS\s+)?([A-Za-z_][A-Za-z0-9_]*))?',
                normalized,
                re.IGNORECASE,
            )
            if match:
                pending_table_keyword = ''
                _, table_name, alias = match.groups()
                display = tables.get(table_name, table_name)
                alias_map[table_name] = display
                if alias:
                    alias_map[alias] = display
                continue

            upper_line = normalized.upper()
            if upper_line in {'FROM', 'INNER JOIN', 'LEFT JOIN', 'RIGHT JOIN', 'FULL JOIN', 'CROSS JOIN', 'JOIN'}:
                pending_table_keyword = upper_line
                continue

            if pending_table_keyword:
                match = re.match(r'^([A-Za-z_][A-Za-z0-9_]*)(?:\s+(?:AS\s+)?([A-Za-z_][A-Za-z0-9_]*))?$', normalized, re.IGNORECASE)
                pending_table_keyword = ''
                if not match:
                    continue
                table_name, alias = match.groups()
                display = tables.get(table_name, table_name)
                alias_map[table_name] = display
                if alias:
                    alias_map[alias] = display
        return alias_map

    def _translate_table_reference(self, text, translation):
        tables = translation['tables']
        text = text.strip()
        match = re.match(r'^([A-Za-z_][A-Za-z0-9_]*)(?:\s+(?:AS\s+)?([A-Za-z_][A-Za-z0-9_]*))?$', text, re.IGNORECASE)
        if not match:
            return self._translate_sql_content(text, translation, {})
        table_name, _alias = match.groups()
        return tables.get(table_name, table_name)

    def _translate_sql_content(self, text, translation, alias_map):
        text = text.strip()
        if not text:
            return ''

        for raw, label in sorted(translation['constants'].items(), key=lambda item: -len(item[0])):
            text = text.replace(raw, label)

        def _param_repl(match):
            param_name = match.group(1)
            return translation['params'].get(param_name, f'[パラメーター.{param_name}]')

        text = re.sub(r'#\{([A-Za-z_][A-Za-z0-9_]*)\}', _param_repl, text)

        def _dotted_repl(match):
            lhs = match.group(1)
            rhs = match.group(2)
            lhs_label = alias_map.get(lhs, translation['tables'].get(lhs, lhs))
            rhs_label = translation['columns'].get(rhs, rhs)
            return f'{lhs_label}.{rhs_label}'

        text = re.sub(r'\b([A-Za-z_][A-Za-z0-9_]*)\.([A-Za-z_][A-Za-z0-9_]*)\b', _dotted_repl, text)

        for table_name, label in sorted(translation['tables'].items(), key=lambda item: -len(item[0])):
            text = re.sub(rf'\b{re.escape(table_name)}\b', label, text)

        for column_name, label in sorted(translation['columns'].items(), key=lambda item: -len(item[0])):
            text = re.sub(rf'\b{re.escape(column_name)}\b', label, text)

        return text

    def _parse_mybatis_sql_clauses(self, sql_text, translation):
        sql_lines = self._split_sql_lines(sql_text)
        alias_map = self._extract_sql_alias_map(sql_lines, translation)
        clauses = []
        current = None

        for raw_line in sql_lines:
            matched_keyword = None
            upper_line = raw_line.upper()
            for keyword in SQL_CLAUSE_KEYWORDS:
                if upper_line == keyword or upper_line.startswith(f'{keyword} '):
                    matched_keyword = keyword
                    break

            if matched_keyword:
                remainder = raw_line[len(matched_keyword):].strip()
                current = {'keyword': matched_keyword, 'lines': []}
                if remainder:
                    if matched_keyword in {'INNER JOIN', 'LEFT JOIN', 'RIGHT JOIN', 'FULL JOIN', 'CROSS JOIN', 'JOIN'} and re.search(r'\s+ON\s+', remainder, re.IGNORECASE):
                        table_part, on_part = re.split(r'\s+ON\s+', remainder, maxsplit=1, flags=re.IGNORECASE)
                        current['lines'].append(self._translate_table_reference(table_part, translation))
                        clauses.append(current)
                        current = {'keyword': 'ON', 'lines': [self._translate_sql_content(on_part, translation, alias_map)]}
                    elif matched_keyword in {'FROM', 'INNER JOIN', 'LEFT JOIN', 'RIGHT JOIN', 'FULL JOIN', 'CROSS JOIN', 'JOIN', 'UPDATE', 'DELETE FROM', 'INSERT INTO'}:
                        current['lines'].append(self._translate_table_reference(remainder, translation))
                    else:
                        current['lines'].append(self._translate_sql_content(remainder, translation, alias_map))
                clauses.append(current)
                continue

            if current is None:
                current = {'keyword': '', 'lines': []}
                clauses.append(current)

            if current.get('keyword') in {'FROM', 'INNER JOIN', 'LEFT JOIN', 'RIGHT JOIN', 'FULL JOIN', 'CROSS JOIN', 'JOIN', 'UPDATE', 'DELETE FROM', 'INSERT INTO'} and not current.get('lines'):
                current['lines'].append(self._translate_table_reference(raw_line, translation))
            else:
                current['lines'].append(self._translate_sql_content(raw_line, translation, alias_map))

        return clauses

    def _build_mybatis_sql_blocks(self, item):
        xml_path = item.get('mapper_xml') or self.config.get('mapper_xml')
        if not xml_path:
            raise ValueError('mybatis_sql requires mapper_xml in item or config')

        statement_ids = item.get('statement_ids') or [item.get('statement_id')]
        statement_ids = [statement_id for statement_id in statement_ids if statement_id]
        if not statement_ids:
            raise ValueError('mybatis_sql requires statement_id or statement_ids')

        translation = self._get_sql_translation_map(item)
        blocks = []
        for statement_id in statement_ids:
            statement = self._load_mybatis_statement(xml_path, statement_id)
            title = item.get('title')
            if not title:
                statement_type = statement.get('statement_type')
                title = {
                    'select': '[取得SQL]',
                    'insert': '[登録SQL]',
                    'update': '[更新SQL]',
                    'delete': '[削除SQL]',
                }.get(statement_type, '[SQL]')

            comment = translation['comments'].get(statement_id) or statement.get('comment', '')
            clauses = self._parse_mybatis_sql_clauses(statement['sql_text'], translation)
            blocks.append({
                'statement_id': statement_id,
                'comment': comment,
                'title': title,
                'clauses': clauses,
            })

        return blocks

    def _extract_db_accesses_from_sql_text(self, sql_text, statement_type, translation):
        sql_lines = self._split_sql_lines(sql_text)
        pending_keyword = ''
        seen = set()
        accesses = []

        def add_access(op, table_name):
            table_label = translation['tables'].get(table_name, table_name)
            key = (op, table_label)
            if not table_label or key in seen:
                return
            seen.add(key)
            accesses.append({'op': op, 'table': table_label})

        for raw_line in sql_lines:
            normalized = re.sub(r'\s+', ' ', raw_line.strip())
            upper_line = normalized.upper()
            matched_keyword = None
            matched_table = None

            for keyword in ('DELETE FROM', 'INSERT INTO', 'UPDATE', 'FROM', 'INNER JOIN', 'LEFT JOIN', 'RIGHT JOIN', 'FULL JOIN', 'CROSS JOIN', 'JOIN'):
                match = re.match(
                    rf'^{re.escape(keyword)}\s+([A-Za-z_][A-Za-z0-9_]*)',
                    normalized,
                    re.IGNORECASE,
                )
                if match:
                    matched_keyword = keyword
                    matched_table = match.group(1)
                    break

            if matched_keyword and matched_table:
                pending_keyword = ''
                if matched_keyword in {'DELETE FROM', 'INSERT INTO', 'UPDATE'}:
                    add_access('w', matched_table)
                else:
                    add_access('r', matched_table)
                continue

            if upper_line in {'DELETE FROM', 'INSERT INTO', 'UPDATE', 'FROM', 'INNER JOIN', 'LEFT JOIN', 'RIGHT JOIN', 'FULL JOIN', 'CROSS JOIN', 'JOIN'}:
                pending_keyword = upper_line
                continue

            if pending_keyword:
                match = re.match(r'^([A-Za-z_][A-Za-z0-9_]*)', normalized, re.IGNORECASE)
                keyword = pending_keyword
                pending_keyword = ''
                if match:
                    table_name = match.group(1)
                    if keyword in {'DELETE FROM', 'INSERT INTO', 'UPDATE'}:
                        add_access('w', table_name)
                    else:
                        add_access('r', table_name)

        if statement_type == 'select':
            return accesses
        return accesses

    def _walk_processing_detail_items(self):
        detail = self.config.get('processing_detail', {}) or {}
        steps = list(detail.get('steps', []))

        def walk_content(items):
            for item in items or []:
                if not isinstance(item, dict):
                    continue
                yield item
                if item.get('type') == 'branch':
                    yield from walk_content(item.get('content', []))

        def walk_steps(nodes):
            for step in nodes or []:
                for item in walk_content(step.get('content', [])):
                    yield item
                for child in walk_steps(step.get('children', [])):
                    yield child

        return list(walk_steps(steps))

    def _derive_db_accesses_from_processing_detail(self):
        accesses = []
        seen = set()

        def add_access(op, table):
            key = (op, table)
            if not table or key in seen:
                return
            seen.add(key)
            accesses.append({'op': op, 'table': table})

        for item in self._walk_processing_detail_items():
            item_type = item.get('type')
            if item_type == 'mybatis_sql':
                xml_path = item.get('mapper_xml') or self.config.get('mapper_xml')
                statement_ids = item.get('statement_ids') or [item.get('statement_id')]
                statement_ids = [statement_id for statement_id in statement_ids if statement_id]
                translation = self._get_sql_translation_map(item)
                for statement_id in statement_ids:
                    statement = self._load_mybatis_statement(xml_path, statement_id)
                    for access in self._extract_db_accesses_from_sql_text(
                        statement.get('sql_text', ''),
                        statement.get('statement_type', ''),
                        translation,
                    ):
                        add_access(access['op'], access['table'])
            elif item_type == 'sql_structured':
                title = item.get('title', '')
                write_keywords = {'[更新SQL]', '[削除SQL]', '[登録SQL]'}
                default_op = 'w' if title in write_keywords else 'r'
                for clause in item.get('clauses', []):
                    keyword = clause.get('keyword', '').upper()
                    if keyword in {'FROM', 'JOIN', 'INNER JOIN', 'LEFT JOIN', 'RIGHT JOIN', 'FULL JOIN', 'CROSS JOIN'}:
                        for line in clause.get('lines', []):
                            add_access('r', str(line).strip())
                    elif keyword in {'DELETE', 'DELETE FROM', 'UPDATE', 'INSERT INTO'}:
                        for line in clause.get('lines', []):
                            add_access('w', str(line).strip())
                    elif keyword and default_op:
                        for line in clause.get('lines', []):
                            if 'テーブル' in str(line):
                                add_access(default_op, str(line).strip())

        return accesses

    # ------------------------------------------------------------------
    # 処理概要 (text + flowchart via DrawingML)
    # ------------------------------------------------------------------
    def fill_overview(self):
        ws = self.wb['処理概要']
        overview = self.config.get('overview', {})

        # Row 7: 機能概要 text (merged cell A7:K49 in template)
        ws['A7'] = overview.get('summary', '')
        ws['A7'].font = FONT_NORMAL
        ws['A7'].alignment = ALIGN_VTL

        # Flowchart will be injected as DrawingML in post-processing
        steps = overview.get('flow_steps', [])
        if steps:
            self._overview_steps = steps
            print(f'  [ok] 処理概要: {len(steps)}ステップ (DrawingML図形で生成予定)')
        else:
            self._overview_steps = []

    # ------------------------------------------------------------------
    # APIシーケンス&DFD
    # ------------------------------------------------------------------
    def fill_sequence(self):
        ws = self.wb['APIシーケンス&DFD']
        seq = self._build_sequence_config()

        # Row 9: Component name
        ws['A9'] = seq.get('client_component', '')
        ws['A9'].font = FONT_NORMAL
        ws['A9'].alignment = ALIGN_VCL

        row = 10
        title = seq.get('api_title', '')
        if title:
            ws[f'C{row}'] = title
            ws[f'C{row}'].font = FONT_NORMAL
            ws[f'C{row}'].alignment = ALIGN_VCL
            row += 1

        steps = seq.get('steps', [])
        for step in steps:
            step_type = step.get('type', 'backend')

            if step_type == 'request':
                # Client → Server communication
                self._seq_request_rows.append(row)  # Track for arrow injection
                step_no = step.get('step_no')
                if step_no not in (None, ''):
                    ws[f'A{row}'] = step_no
                    ws[f'A{row}'].font = FONT_NORMAL
                    ws[f'A{row}'].alignment = ALIGN_VCC
                ws[f'B{row}'] = step.get('trigger', '')
                ws[f'B{row}'].font = FONT_NORMAL
                ws[f'B{row}'].alignment = ALIGN_VCL
                ws[f'C{row}'] = step.get('communication', step.get('url', ''))
                ws[f'C{row}'].font = FONT_NORMAL
                ws[f'C{row}'].alignment = ALIGN_VCL
                description = step.get('description', '')
                if description:
                    ws[f'E{row}'] = description
                    ws[f'E{row}'].font = FONT_NORMAL
                    ws[f'E{row}'].alignment = ALIGN_VCL

                db_accesses = self._expand_sequence_db_accesses(step)
                if db_accesses:
                    first_access = db_accesses[0]
                    ws[f'G{row}'] = first_access.get('op', '')
                    ws[f'G{row}'].font = FONT_NORMAL
                    ws[f'G{row}'].alignment = ALIGN_VCL
                    self._write_detail_text_cell(
                        ws,
                        f'H{row}',
                        self._format_sequence_table_name(first_access.get('table', '')),
                    )

                params = step.get('params', [])
                details = step.get('details', [])
                extra_rows = max(len(params), len(details), max(len(db_accesses) - 1, 0))
                for extra_idx in range(extra_rows):
                    row += 1
                    if extra_idx < len(params):
                        ws[f'C{row}'] = f'・{params[extra_idx]}'
                        ws[f'C{row}'].font = FONT_NORMAL
                        ws[f'C{row}'].alignment = ALIGN_VCL
                    if extra_idx < len(details):
                        ws[f'E{row}'] = f'　{details[extra_idx]}'
                        ws[f'E{row}'].font = FONT_NORMAL
                        ws[f'E{row}'].alignment = ALIGN_VCL
                    db_idx = extra_idx + 1
                    if db_idx < len(db_accesses):
                        ws[f'G{row}'] = db_accesses[db_idx].get('op', '')
                        ws[f'G{row}'].font = FONT_NORMAL
                        ws[f'G{row}'].alignment = ALIGN_VCL
                        self._write_detail_text_cell(
                            ws,
                            f'H{row}',
                            self._format_sequence_table_name(db_accesses[db_idx].get('table', '')),
                        )
                row += 1

            elif step_type == 'backend':
                step_no = step.get('step_no')
                if step_no not in (None, ''):
                    ws[f'A{row}'] = step_no
                    ws[f'A{row}'].font = FONT_NORMAL
                    ws[f'A{row}'].alignment = ALIGN_VCC
                ws[f'E{row}'] = step.get('description', '')
                ws[f'E{row}'].font = FONT_NORMAL
                ws[f'E{row}'].alignment = ALIGN_VCL
                db_accesses = self._expand_sequence_db_accesses(step)
                if db_accesses:
                    first_access = db_accesses[0]
                    ws[f'G{row}'] = first_access.get('op', '')
                    ws[f'G{row}'].font = FONT_NORMAL
                    ws[f'G{row}'].alignment = ALIGN_VCL
                    self._write_detail_text_cell(
                        ws,
                        f'H{row}',
                        self._format_sequence_table_name(first_access.get('table', '')),
                    )

                details = step.get('details', [])
                extra_rows = max(len(details), max(len(db_accesses) - 1, 0))
                for extra_idx in range(extra_rows):
                    row += 1
                    if extra_idx < len(details):
                        ws[f'E{row}'] = f'　{details[extra_idx]}'
                        ws[f'E{row}'].font = FONT_NORMAL
                        ws[f'E{row}'].alignment = ALIGN_VCL
                    db_idx = extra_idx + 1
                    if db_idx < len(db_accesses):
                        ws[f'G{row}'] = db_accesses[db_idx].get('op', '')
                        ws[f'G{row}'].font = FONT_NORMAL
                        ws[f'G{row}'].alignment = ALIGN_VCL
                        self._write_detail_text_cell(
                            ws,
                            f'H{row}',
                            self._format_sequence_table_name(db_accesses[db_idx].get('table', '')),
                        )
                row += 1

            elif step_type == 'response':
                self._seq_response_rows.append(row)  # Track for arrow injection
                step_no = step.get('step_no')
                if step_no not in (None, ''):
                    ws[f'A{row}'] = step_no
                    ws[f'A{row}'].font = FONT_NORMAL
                    ws[f'A{row}'].alignment = ALIGN_VCC
                trigger = step.get('trigger', '')
                if trigger:
                    ws[f'B{row}'] = trigger
                    ws[f'B{row}'].font = FONT_NORMAL
                    ws[f'B{row}'].alignment = ALIGN_VCL
                ws[f'C{row}'] = step.get('status', 'HTTPステータスコード:200')
                ws[f'C{row}'].font = FONT_NORMAL
                ws[f'C{row}'].alignment = ALIGN_VCL
                ws[f'E{row}'] = step.get('description', '　レスポンス返却')
                ws[f'E{row}'].font = FONT_NORMAL
                ws[f'E{row}'].alignment = ALIGN_VCL
                for param in step.get('params', []):
                    row += 1
                    ws[f'C{row}'] = f'・{param}'
                    ws[f'C{row}'].font = FONT_NORMAL
                    ws[f'C{row}'].alignment = ALIGN_VCL
                row += 1

        print(f'  [ok] APIシーケンス&DFD: {len(steps)}ステップ, {row - 10}行')

    def _expand_sequence_db_accesses(self, step):
        db_accesses = list(step.get('db_accesses', []) or [])
        if db_accesses:
            return db_accesses

        db_op = step.get('db_op', '')
        db_table = step.get('db_table', '')
        if db_op in (None, '') or db_table in (None, ''):
            return []

        tables = [part.strip() for part in str(db_table).split('/') if part.strip()]
        if not tables:
            return []
        return [{'op': db_op, 'table': table} for table in tables]

    def _build_sequence_config(self):
        seq = dict(self.config.get('sequence', {}) or {})
        if seq.get('steps'):
            seq.setdefault('client_component', self._infer_client_component())
            seq.setdefault('api_title', self._infer_sequence_title())
            return seq

        seq['client_component'] = seq.get('client_component') or self._infer_client_component()
        seq['api_title'] = seq.get('api_title') or self._infer_sequence_title()
        seq['steps'] = self._build_default_sequence_steps(seq)
        return seq

    def _infer_client_component(self):
        seq = self.config.get('sequence', {}) or {}
        if seq.get('client_component'):
            return seq['client_component']

        for candidate in (
            self.config.get('screen', {}).get('name') if self.config.get('screen') else '',
            self.config.get('cover', {}).get('screen_name') if self.config.get('cover') else '',
            self.config.get('api_info', {}).get('client_component') if self.config.get('api_info') else '',
        ):
            if candidate:
                return candidate

        description = self.config.get('api_info', {}).get('description', '')
        m = re.search(r'([^\s、。]+画面)', description)
        if m:
            return m.group(1)
        return self.config.get('cover', {}).get('api_name', '')

    def _infer_sequence_title(self):
        seq = self.config.get('sequence', {}) or {}
        if seq.get('api_title'):
            return seq['api_title']

        api_name = self.config.get('cover', {}).get('api_name', '')
        if api_name:
            return f'{api_name}API'
        description = self.config.get('api_info', {}).get('description', '')
        if description:
            return f'{description}API'
        return 'API処理'

    def _build_default_sequence_steps(self, seq):
        api = self.config.get('api_info', {})
        url = api.get('url', '')
        response_code = seq.get('response_code', 200)
        derived_db_accesses = self._derive_db_accesses_from_processing_detail()
        request_params = [
            self._pick_param_value(p, 'name', 'param_name', 'field_name', 'item_name')
            for p in self._flatten_params(self.config.get('request_params', []))
            if p.get('depth', 0) == 0
        ]
        response_params = [
            self._pick_param_value(p, 'name', 'param_name', 'field_name', 'item_name')
            for p in self._flatten_params(self.config.get('response_params', []))
            if p.get('depth', 0) == 0
        ]
        backend_desc = (
            seq.get('backend_summary')
            or api.get('description')
            or self.config.get('overview', {}).get('summary', '')
        )
        db_accesses = derived_db_accesses or seq.get('db_accesses', [])
        return [
            {
                'type': 'request',
                'step_no': 1,
                'trigger': seq.get('request_trigger', self._infer_request_trigger(url, backend_desc)),
                'communication': seq.get('request_communication', self._infer_request_operation(url)),
                'description': backend_desc,
                'params': request_params,
                'details': seq.get('request_details', []),
                'db_accesses': db_accesses,
            },
            {
                'type': 'response',
                'trigger': seq.get('response_trigger', self._infer_response_trigger(url)),
                'status': seq.get('response_status', f'httpレスポンス:{response_code}'),
                'description': seq.get('response_description', 'レスポンス返却'),
                'params': response_params,
            },
        ]

    def _infer_request_operation(self, url):
        parts = [part for part in url.strip('/').split('/') if part]
        if not parts:
            return 'request'
        if len(parts) >= 2:
            return '/'.join(parts[-2:])
        return parts[-1]

    def _infer_request_trigger(self, url, description=''):
        key = f'{url} {description}'.lower()
        if 'delete' in key or '削除' in description:
            return '削除実行'
        if 'import' in key or '取込' in description:
            return '取込実行'
        if 'show' in key or 'initial' in key or '初期' in description:
            return '初期表示'
        if 'search' in key or 'list' in key or '一覧' in description:
            return '一覧取得'
        if 'save' in key or 'update' in key or '登録' in description or '更新' in description:
            return '登録・更新'
        return 'API呼出'

    def _infer_response_trigger(self, url):
        key = url.lower()
        if 'delete' in key:
            return '一覧再表示'
        if 'show' in key:
            return '画面表示'
        if 'import' in key:
            return '取込結果反映'
        return '処理結果反映'

    # ------------------------------------------------------------------
    # 処理詳細 (★ v2: wrap_text=False でテキスト溢れ表示)
    # ------------------------------------------------------------------
    def fill_processing_detail(self):
        ws = self.wb['処理詳細']
        detail = self.config.get('processing_detail', {})
        steps = detail.get('steps', [])

        if not steps:
            print('  - 処理詳細: データなし（スキップ）')
            return

        row = 9  # Start after header
        row = self._write_detail_steps(ws, steps, row, depth=0, parent_num='')
        print(f'  [ok] 処理詳細: {row - 9}行生成')

    def _write_detail_steps(self, ws, steps, row, depth, parent_num):
        """Recursively write processing detail steps.

        ★ v2: wrap_text=False — テキストはセルの右端を超えて表示される
              （原本と同じ振る舞い）
        """
        for idx, step in enumerate(steps):
            num = step.get('number', '')
            if not num:
                num = f'{parent_num}{idx + 1}' if not parent_num else f'{parent_num}-{idx + 1}'

            title = step.get('title', '')
            num_col = get_column_letter(DETAIL_BASE_COL + depth)          # number column
            content_col = get_column_letter(DETAIL_BASE_COL + depth + 1)  # title column

            # Write step number and title
            ws[f'{num_col}{row}'] = num
            ws[f'{num_col}{row}'].font = FONT_NORMAL
            self._write_detail_text_cell(ws, f'{content_col}{row}', title)
            row += 1

            # blank line after title
            row += 1

            # Write content items
            content = step.get('content', [])
            row = self._write_content_items(ws, content, row, depth + 1)

            # Recurse children
            children = step.get('children', [])
            if children:
                row = self._write_detail_steps(ws, children, row, depth + 1, num)

            row += 1  # blank line between steps

        return row

    def _write_content_items(self, ws, items, row, depth):
        """Write content items with text overflow (wrap_text=False)."""
        base_col_num = DETAIL_BASE_COL + depth + 1  # content starts one deeper

        for item in items:
            if isinstance(item, str):
                col = get_column_letter(base_col_num)
                self._write_detail_text_cell(ws, f'{col}{row}', item)
                row += 1

            elif isinstance(item, dict):
                item_type = item.get('type', 'text')

                if item_type == 'sub':
                    col = get_column_letter(base_col_num)
                    self._write_detail_text_cell(ws, f'{col}{row}', f"◎{item.get('text', '')}")
                    row += 1

                elif item_type == 'call':
                    col = get_column_letter(base_col_num)
                    self._write_detail_text_cell(ws, f'{col}{row}', item.get('method', ''))
                    row += 1
                    param_col = get_column_letter(base_col_num + 1)
                    for p in item.get('params', []):
                        self._write_detail_text_cell(ws, f'{param_col}{row}', f'・{p}')
                        row += 1

                elif item_type == 'sql':
                    col = get_column_letter(base_col_num)
                    self._write_detail_text_cell(ws, f'{col}{row}', '[取得SQL]')
                    row += 1
                    sql_col = get_column_letter(base_col_num + 1)
                    for line in item.get('lines', []):
                        self._write_detail_text_cell(ws, f'{sql_col}{row}', line)
                        row += 1
                    row += 1  # blank after SQL

                elif item_type == 'sql_structured':
                    title_col = get_column_letter(base_col_num)
                    keyword_col = get_column_letter(base_col_num + 1)
                    detail_col = get_column_letter(base_col_num + 2)

                    self._write_detail_text_cell(ws, f'{title_col}{row}', item.get('title', '[取得SQL]'))
                    row += 1

                    for clause in item.get('clauses', []):
                        keyword = clause.get('keyword', '')
                        if keyword:
                            self._write_detail_text_cell(ws, f'{keyword_col}{row}', keyword)
                            row += 1

                        lines = clause.get('lines', [])
                        for line in lines:
                            self._write_detail_text_cell(ws, f'{detail_col}{row}', line)
                            row += 1

                    row += 1

                elif item_type == 'mybatis_sql':
                    title_col = get_column_letter(base_col_num)
                    keyword_col = get_column_letter(base_col_num + 1)
                    detail_col = get_column_letter(base_col_num + 2)

                    for block in self._build_mybatis_sql_blocks(item):
                        if block.get('comment'):
                            self._write_detail_text_cell(ws, f'{title_col}{row}', block['comment'])
                            row += 1

                        self._write_detail_text_cell(ws, f'{title_col}{row}', block.get('title', '[SQL]'))
                        row += 1

                        for clause in block.get('clauses', []):
                            keyword = clause.get('keyword', '')
                            if keyword:
                                self._write_detail_text_cell(ws, f'{keyword_col}{row}', keyword)
                                row += 1

                            for line in clause.get('lines', []):
                                self._write_detail_text_cell(ws, f'{detail_col}{row}', line)
                                row += 1

                        row += 1

                elif item_type == 'branch':
                    arrow_col = get_column_letter(base_col_num)
                    text_col = get_column_letter(base_col_num + 1)
                    cond = item.get('condition', '')
                    action = item.get('action', '')
                    self._write_detail_text_cell(ws, f'{arrow_col}{row}', '⇒', horizontal='center')
                    branch_text = f'{cond}、{action}' if cond and action else (cond or action)
                    self._write_detail_text_cell(ws, f'{text_col}{row}', branch_text)
                    row += 1
                    # Nested content under branch
                    branch_content = item.get('content', [])
                    if branch_content:
                        row = self._write_content_items(ws, branch_content, row, depth + 1)

                elif item_type == 'data':
                    col = get_column_letter(base_col_num)
                    self._write_detail_text_cell(ws, f'{col}{row}', '[data]')
                    row += 1
                    row = self._write_assignments(ws, item.get('assignments', []),
                                                  row, base_col_num + 1)

                elif item_type == 'text':
                    col = get_column_letter(base_col_num)
                    self._write_detail_text_cell(ws, f'{col}{row}', item.get('text', ''))
                    row += 1

        return row

    def _write_assignments(self, ws, assignments, row, param_col_num):
        """Write [data] block assignments: param ← source."""
        for a in assignments:
            param = a.get('param', '')
            source = a.get('source', '')

            pcol = get_column_letter(param_col_num)
            arrow_col = get_column_letter(param_col_num + ARROW_OFFSET)
            val_col = get_column_letter(param_col_num + ARROW_OFFSET + 1)

            self._write_detail_text_cell(ws, f'{pcol}{row}', param)
            if source:
                self._write_detail_text_cell(ws, f'{arrow_col}{row}', '←', horizontal='center')
                self._write_detail_text_cell(ws, f'{val_col}{row}', source)
            row += 1

            # Nested children (deeper params)
            children = a.get('children', [])
            if children:
                row = self._write_assignments(ws, children, row, param_col_num + 1)

        return row

    # ------------------------------------------------------------------
    # 改定履歴
    # ------------------------------------------------------------------
    def fill_revision_history(self):
        ws = self.wb['改定履歴']
        cover = self.config.get('cover', {})
        ws['D8'] = cover.get('create_date', '')
        ws['E8'] = '新規作成'
        ws['K8'] = cover.get('author', '')

    # ------------------------------------------------------------------
    # Generate
    # ------------------------------------------------------------------
    def generate(self, output_path: str):
        """Run all generators and save."""
        print('=== API設計書 生成開始 (v2) ===')

        self.fill_cover()
        print('  [ok] 表紙')

        self.fill_request_params()
        self.fill_response_params()
        self.fill_overview()
        self.fill_sequence()
        self.fill_processing_detail()
        self.fill_revision_history()
        print('  [ok] 改定履歴')

        try:
            # Save workbook first
            self.wb.save(output_path)
            print(f'  [ok] ワークブック保存: {output_path}')

            # Post-process: inject DrawingML shapes
            drawings_to_inject = {}

            # 処理概要 flowchart
            if hasattr(self, '_overview_steps') and self._overview_steps:
                drawings_to_inject['処理概要'] = _make_flowchart_xml(self._overview_steps)

            # APIシーケンス&DFD arrows
            if self._seq_request_rows or self._seq_response_rows:
                drawings_to_inject['APIシーケンス&DFD'] = _make_sequence_arrows_xml(
                    self._seq_request_rows, self._seq_response_rows)

            if drawings_to_inject:
                inject_drawings_into_xlsx(output_path, drawings_to_inject)

            print(f'\n[ok] 完了: {output_path}')
        finally:
            self.wb.close()


# ===================================================================
# CLI
# ===================================================================
def merge_project_config(config, project_config):
    """Merge project-level config into endpoint config.

    Project config provides defaults for:
      - cover.company ← project_config.company.name
      - cover.project ← project_config.project.name
      - cover.system  ← project_config.project.system_name
      - cover.author  ← project_config.author.name
      - cover.create_date ← today (if not set)
      - cover.update_date ← today (if not set)
      - cover.update_author ← cover.author (if not set)

    Endpoint config values always take precedence over project defaults.
    """
    if not project_config:
        return config

    cover = config.setdefault('cover', {})

    # Company
    company = project_config.get('company', {})
    if not cover.get('company') and company.get('name'):
        cover['company'] = company['name']

    # Project
    project = project_config.get('project', {})
    if not cover.get('project') and project.get('name'):
        cover['project'] = project['name']
    if not cover.get('system') and project.get('system_name'):
        cover['system'] = project['system_name']

    # Author
    author = project_config.get('author', {})
    default_author = author.get('name') or 'isi'
    if not cover.get('author'):
        cover['author'] = default_author

    # Dates: auto-fill today if not specified
    today = datetime.now().strftime('%Y-%m-%d')
    if not cover.get('create_date'):
        cover['create_date'] = today
    if not cover.get('update_date'):
        cover['update_date'] = today
    if not cover.get('update_author'):
        cover['update_author'] = cover.get('author', default_author)

    return config


def main():
    parser = argparse.ArgumentParser(description='API設計書ジェネレーター v2')
    parser.add_argument('config', help='Config JSON path')
    parser.add_argument('-p', '--project-config', default=None,
                        help='Project config JSON path (shared metadata)')
    parser.add_argument('-t', '--template', default=None,
                        help='Template xlsx path')
    parser.add_argument('-o', '--output', default=None,
                        help='Output xlsx path')
    args = parser.parse_args()

    config = json.loads(Path(args.config).read_text(encoding='utf-8'))

    # Load and merge project config
    project_config = None
    pc_path = args.project_config
    if not pc_path:
        # Auto-detect project_config.json in configs/ directory
        auto_path = Path(__file__).parent.parent / 'configs' / 'project_config.json'
        if auto_path.exists():
            pc_path = str(auto_path)

    if pc_path:
        project_config = json.loads(Path(pc_path).read_text(encoding='utf-8'))
        config = merge_project_config(config, project_config)
        print(f'  [ok] プロジェクト設定読み込み: {pc_path}')

    validate_japanese_content(config)
    print('  [ok] 日本語必須チェック')

    # Resolve template path
    template = args.template
    if not template:
        if project_config and project_config.get('template_paths', {}).get('api_spec'):
            template = str(Path(__file__).parent.parent / project_config['template_paths']['api_spec'])
        else:
            template = str(Path(__file__).parent.parent / 'assets' / 'api_template_clean.xlsx')

    output = args.output or build_default_output_name(config)

    gen = ApiSpecGenerator(config, template)
    gen.generate(output)


if __name__ == '__main__':
    main()
